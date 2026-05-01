#!/Users/amromer/msd_env/bin/python
#HOW TO RUN 
# python3 msd_4pl_analysis.py --msd 185-008_ControlVariability/23N3QAYE23_2026-04-01-082642.txt --platemap 185-008_ControlVariability/AssayPlateMap.csv --output 185-008_ControlVariability/results.xlsx
"""
MSD 4PL Analysis Tool
=====================
Parses MSD instrument .txt files (1, 4, or 10 spots per well; multi-plate),
fits 4-parameter logistic regression per analyte/spot, interpolates unknown
concentrations, and outputs a formatted Excel workbook with standard curves.

USAGE
-----
  Interactive mode (single-page GUI):
    python3 msd_4pl_analysis.py
    python3 msd_4pl_analysis.py --gui

  Command-line mode:
    python3 msd_4pl_analysis.py --msd <data.txt> --platemap <map.csv> --output <results.xlsx>
    python3 msd_4pl_analysis.py --msd <data.txt> --platemap <map.csv> --output <results.xlsx> --spots 4 --units pg/mL --cv-threshold 30 --lloq-method 3xblank --dilution-factors 1,2,1

  Running with no arguments or with --gui opens a single-page GUI to configure
  all options and select files.

  --msd              MSD instrument .txt data file (supports multi-plate files)
  --platemap         Plate map CSV in grid format (see below)
  --output           Output Excel file path (default: msd_4pl_results.xlsx)
  --spots            Override spots per well: 1, 4, or 10 (auto-detected if omitted)
  --units            Optional units string to append to interpolated concentration headers
  --cv-threshold     Optional %CV threshold for All Unknowns highlight (default 25)
  --lloq-method      LLOQ calculation method: 'current' (mean+10*SD) or '3xblank' (3x blank mean)
  --dilution-factors Optional per-plate dilution factors as comma-separated values (e.g. 1,2,1)
  --gui              Force interactive GUI mode
  --rerun            Rerun the last analysis with saved parameters

PLATE MAP FORMAT
----------------
The plate map is a CSV in 96-well grid layout. Row letters (A-H) are the first
column; column numbers (1-12) are the header row.

  Example:
    ,1,2,3,4,5,6,7,8,9,10,11,12
    A,800000,800000,fCtx,mCtx,Cd,Put,Hp,fCtx,mCtx,Cd,Put,Hp
    B,200000,200000,fCtx,mCtx,Cd,Put,Hp,fCtx,mCtx,Cd,Put,Hp
    ...
    H,Buffer Only,Buffer Only,HQC,MQC,LQC,,,,,,,

CELL CLASSIFICATION RULES
--------------------------
Each cell in the plate map is classified automatically based on its content:

  Standard   — Cell contains a PURELY NUMERIC value (integer or decimal).
               The number is used as the known concentration for curve fitting.
               Commas are stripped before parsing (e.g. "3,125" → 3125).
               Examples: 800000, 781.25, 3125, 0.61

  Unknown    — Cell contains ANY TEXT or a MIX of text and numbers.
               Treated as an unknown sample; the cell value becomes the sample name.
               The concentration will be interpolated from the fitted 4PL curve.
               Examples: fCtx, mCtx, HQC, Sample_3, STD-1, 800000 pg/ml

  Blank      — Cell matches one of these keywords (case-insensitive):
               "Buffer Only", "Blank", "Buffer", "BG", "Background", "0"
               Blanks are included in the curve fit at concentration = 0 and used
               to calculate LLOQ (mean + 10 × SD of blank signals).

  Empty      — Cell is empty or contains only whitespace. Skipped entirely.

  NOTE: There is no requirement for standards or samples to be in specific
  wells or orientations. Standards can be scattered anywhere on the plate.
  The only requirement is at least 4 unique standard concentrations per
  curve for a successful 4PL fit.

  CAUTION: A sample name that is purely numeric (e.g. a sample ID "12345")
  will be misclassified as a standard at concentration 12345. To avoid this,
  include at least one non-numeric character in sample names (e.g. "S-12345").

GROUP PREFIX (MULTIPLE CURVES PER PLATE)
----------------------------------------
To run multiple independent standard curves on the same plate, prefix cell
values with a group tag followed by a colon:

    GroupName:value

  Examples:
    CurveA:800000      → Standard at 800000, assigned to group "CurveA"
    CurveA:fCtx        → Unknown sample "fCtx", assigned to group "CurveA"
    CurveB:500000      → Standard at 500000, assigned to group "CurveB"
    CurveB:SampleX     → Unknown sample "SampleX", assigned to group "CurveB"

  Each group gets its own independent 4PL fit, LLOQ/ULOQ, and Excel sheet.
  Blanks WITHOUT a group prefix are shared across all groups automatically.
  Cells without any prefix belong to a single default group (backward compatible).

  NOTE: The group prefix can be up to 20 characters. The colon ":" is the
  delimiter, so avoid colons in sample names unless using the group feature.
  A prefix like "1:2" would be interpreted as group "1", value "2".

MULTI-PLATE SUPPORT
-------------------
Multiple plate maps can be stacked vertically in a single CSV, separated by
one or more blank rows. Each block is assigned a plate number (1, 2, 3, ...)
in order and matched to the corresponding plate in the MSD file.

  Example (two plates):
    ,1,2,...,12
    A,800000,...
    ...H,...
                        ← blank row separates plates
    ,1,2,...,12         ← optional repeated header
    A,500000,...
    ...H,...

  If only one plate map is provided but the MSD file contains multiple plates,
  the single map is reused for all plates.

OUTPUT
------
The Excel workbook contains:
  - Summary sheet: 4PL parameters, LLOQ, R² for all spots/groups, plus an
    overlay chart showing all fitted curves on one plot.
  - Per-spot sheets: Detailed standard curve data, blanks, interpolated
    unknowns with ULOQ/LLOQ flags, and an MSD-style log-log chart with
    detection range bands.
  - All Unknowns sheet: Consolidated table of all unknown samples grouped
    by sample name, with averaged signals and concentrations.
"""

import re, sys, argparse, os, tempfile, json, subprocess, platform, functools

LAST_RUN_PATH = os.path.join(os.path.expanduser('~'), '.msd_4pl_last_run.json')
MAX_RUN_HISTORY = 5

# ── Analysis thresholds ───────────────────────────────────────────────────────
R2_GOOD         = 0.99   # R² ≥ this → "Good" curve fit
R2_ACCEPTABLE   = 0.95   # R² ≥ this → "Acceptable" curve fit (else "Poor")
QC_RECOVERY_LOW  = 70.0  # % recovery below this → QC failure
QC_RECOVERY_HIGH = 130.0 # % recovery above this → QC failure
DEFAULT_CV_THRESHOLD = 25.0  # %CV above this → flagged

def _load_run_history():
    """Return list of up to MAX_RUN_HISTORY prior run dicts, newest first.
    Handles legacy single-dict format transparently."""
    try:
        with open(LAST_RUN_PATH, 'r') as f:
            data = json.load(f)
        if isinstance(data, dict):          # legacy single-entry
            return [data]
        return data[:MAX_RUN_HISTORY]
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def _save_run_to_history(entry):
    """Prepend entry to the run history list and trim to MAX_RUN_HISTORY.
    If an entry with the same (msd, platemap, output) already exists it is
    replaced rather than duplicated — re-runs update in place."""
    from datetime import datetime
    entry.setdefault('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M'))
    history = _load_run_history()
    # Remove any existing entry for the same experiment
    key = (entry.get('msd'), entry.get('platemap'), entry.get('output'))
    history = [h for h in history
               if (h.get('msd'), h.get('platemap'), h.get('output')) != key]
    history.insert(0, entry)
    history = history[:MAX_RUN_HISTORY]
    with open(LAST_RUN_PATH, 'w') as f:
        json.dump(history, f, indent=2)

def _run_label(entry):
    """Short human-readable label for a run history entry."""
    status = entry.get('status', '')
    icon = '✓' if status == 'pass' else ('✗' if status == 'fail' else ' ')
    ts  = entry.get('timestamp', '')
    msd = os.path.basename(entry.get('msd') or '') or '—'
    out = os.path.basename(entry.get('output') or '') or '—'
    return f"{icon}  {ts}  |  {msd}  →  {out}"

from io import StringIO
from collections import defaultdict

def _ensure_deps():
    """Lazy-load all heavy analysis dependencies the first time an analysis runs.
    Keeps GUI startup near-instant (only stdlib loads at launch)."""
    if 'np' in globals():
        return
    g = globals()
    try:
        import numpy as np;          g['np'] = np
        import pandas as pd;         g['pd'] = pd
        from scipy.optimize import curve_fit; g['curve_fit'] = curve_fit
        from openpyxl import Workbook; g['Workbook'] = Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        g.update(Font=Font, PatternFill=PatternFill, Alignment=Alignment,
                 Border=Border, Side=Side)
        from openpyxl.chart import ScatterChart, Reference, Series
        g.update(ScatterChart=ScatterChart, Reference=Reference, Series=Series)
        from openpyxl.drawing.image import Image as XlImage; g['XlImage'] = XlImage
        from openpyxl.utils import get_column_letter; g['get_column_letter'] = get_column_letter
        import matplotlib; matplotlib.use('Agg'); g['matplotlib'] = matplotlib
        import matplotlib.pyplot as plt;   g['plt'] = plt
        import matplotlib.ticker as ticker; g['ticker'] = ticker
        import warnings; warnings.filterwarnings('ignore')
    except ModuleNotFoundError as e:
        missing = str(e).split("'")[1] if "'" in str(e) else str(e)
        msg = (f"Missing required package: {missing}\n"
               f"Install with: python3 -m pip install numpy pandas scipy openpyxl matplotlib")
        print(msg)
        raise RuntimeError(msg) from e

    # Openpyxl style constants (constructed once, reused across all sheets)
    g['HEADER_FILL'] = PatternFill('solid', fgColor='2F5496')
    g['HEADER_FONT'] = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    g['DATA_FONT']   = Font(name='Arial', size=10)
    g['BOLD_FONT']   = Font(bold=True, name='Arial', size=10)
    g['SECTION_FONT']= Font(bold=True, name='Arial', size=12, color='2F5496')
    g['THIN_BORDER'] = Border(
        left=Side('thin', color='B4B4B4'), right=Side('thin', color='B4B4B4'),
        top=Side('thin', color='B4B4B4'), bottom=Side('thin', color='B4B4B4'))
    g['STD_FILL']     = PatternFill('solid', fgColor='E2EFDA')
    g['UNK_FILL']     = PatternFill('solid', fgColor='FFF2CC')
    g['BLANK_FILL']   = PatternFill('solid', fgColor='F2F2F2')
    g['CV_GOOD_FILL'] = PatternFill('solid', fgColor='D9EAD3')
    g['CV_BAD_FILL']  = PatternFill('solid', fgColor='F8CBAD')
    g['PASS_FONT']    = Font(name='Arial', size=10, color='006100')
    g['WARN_FONT']    = Font(name='Arial', size=10, color='9C5700')
    g['FAIL_FONT']    = Font(name='Arial', size=10, color='9C0006')


# ═══════════════════════════════════════════════════════════════════════════════
# 4PL MODEL
# ═══════════════════════════════════════════════════════════════════════════════

QC_LEVELS = ["ULOQ", "HQC", "MQC", "LQC", "LLOQ"]

@functools.lru_cache(maxsize=None)
def _identify_qc_level(sample_name):
    """Return which QC level (ULOQ/HQC/MQC/LQC/LLOQ) this sample represents, or None."""
    upper = sample_name.upper()
    for level in QC_LEVELS:
        if level in upper:
            return level
    return None


def four_pl(x, a, b, c, d):
    """4PL: a=min asymptote, b=Hill slope, c=inflection (EC50), d=max asymptote"""
    return d + (a - d) / (1.0 + (x / c) ** b)

def inverse_4pl(y, a, b, c, d):
    """Solve 4PL for x given y."""
    if abs(b) < 1e-9:          # degenerate Hill slope → undefined inverse
        return np.nan
    denom = y - d
    if denom == 0:
        return np.nan
    ratio = (a - d) / denom - 1.0
    if ratio <= 0:
        return np.nan
    return c * (ratio ** (1.0 / b))

def fit_4pl(conc, signal):
    """
    Fit 4PL model with 1/y² weighted least-squares.

    Weighting: σ_i = y_i so the optimiser minimises Σ[(y_i − f(x_i)) / y_i]²,
    i.e. it minimises relative (percentage) residuals — the correct criterion for
    assay signals that span orders of magnitude with roughly constant CV.
    absolute_sigma=False means σ values define relative importance only and are not
    assumed to be true measurement standard deviations.

    Blanks (conc=0) are included; they anchor the lower asymptote (a parameter).
    At x=0 the model returns a, so including blanks constrains a ≈ blank signal.

    R² is computed on the same 1/y² weighted scale as the fit.
    Falls back to unweighted fit if the weighted optimisation fails.

    Returns (popt, weighted_r2) or (None, None).
    """
    conc  = np.asarray(conc,   float)
    signal = np.asarray(signal, float)

    # Require: non-negative concentration, finite & strictly positive signal
    # (signal > 0 required for 1/y² weights; MSD blanks always have positive signal)
    mask = (conc >= 0) & np.isfinite(signal) & (signal > 0)
    c_fit = conc[mask]
    s_fit = signal[mask]

    if len(c_fit) < 4:
        return None, None

    # Initial parameter guesses
    a0 = float(np.min(s_fit))                          # lower asymptote
    d0 = float(np.max(s_fit))                          # upper asymptote
    pos = c_fit[c_fit > 0]
    # Geometric mean of positive concentrations → midpoint on log scale (better than median)
    c0 = float(np.exp(np.mean(np.log(pos)))) if len(pos) > 0 else 1.0
    b0 = 1.0                                           # Hill slope

    # Hill slope bounded to physically meaningful range; c bounded > 0
    bounds = ([-np.inf, 0.01, 1e-15, -np.inf],
              [ np.inf, 20.0,  np.inf,  np.inf])

    def _weighted_r2(params):
        """1/y² weighted R²: consistent with the fitting criterion."""
        y_pred  = four_pl(c_fit, *params)
        w       = 1.0 / s_fit ** 2
        y_wmean = np.average(s_fit, weights=w)
        ss_res  = np.sum(w * (s_fit - y_pred)  ** 2)
        ss_tot  = np.sum(w * (s_fit - y_wmean) ** 2)
        return float(1.0 - ss_res / ss_tot) if ss_tot > 0 else 0.0

    # ── Primary: 1/y² weighted fit ──────────────────────────────────────────
    # σ_i = y_i → minimises Σ(relative_residual²)
    try:
        popt, _ = curve_fit(
            four_pl, c_fit, s_fit,
            p0=[a0, b0, c0, d0],
            sigma=s_fit,            # 1/y² weighting
            absolute_sigma=False,   # σ defines relative weights, not true std-devs
            maxfev=5000,
            bounds=bounds,
        )
        return popt, _weighted_r2(popt)
    except (RuntimeError, ValueError):
        pass  # RuntimeError = maxfev exceeded or no convergence; ValueError = bad input

    # ── Fallback: unweighted fit ─────────────────────────────────────────────
    try:
        popt, _ = curve_fit(
            four_pl, c_fit, s_fit,
            p0=[a0, b0, c0, d0],
            maxfev=5000,
            bounds=bounds,
        )
        return popt, _weighted_r2(popt)
    except (RuntimeError, ValueError):
        return None, None


# ═══════════════════════════════════════════════════════════════════════════════
# CHART GENERATION (MSD Discovery Workbench style)
# ═══════════════════════════════════════════════════════════════════════════════

def generate_std_curve_chart(res, tmp_dir, lloq_method='current'):
    """
    Generate a log-log standard curve plot matching MSD Discovery Workbench style.
    Returns path to saved PNG image, or None if curve fit failed.
    """
    params = res.get('params')
    standards = res.get('standards', [])
    blanks = res.get('blanks', [])
    if params is None or not standards:
        return None

    a, b, c, d = params
    plate = res['plate']
    spot = res['spot']
    group = res.get('group', '')

    # Collect standard data
    std_concs = np.array([s['conc'] for s in standards if s['conc'] > 0])
    std_sigs = np.array([s['signal'] for s in standards if s['conc'] > 0])
    if len(std_concs) == 0:
        return None

    # Calculate LLOQ signal from blanks
    lloq_sig = None
    if blanks:
        bsigs = [bl['signal'] for bl in blanks if np.isfinite(bl['signal'])]
        lloq_sig = calculate_lloq_signal(bsigs, lloq_method)

    # ULOQ signal = fitted signal at highest standard concentration
    uloq_conc = np.max(std_concs)
    uloq_sig = four_pl(uloq_conc, *params)

    # LLOQ concentration from signal
    lloq_conc = None
    if lloq_sig is not None:
        try:
            lloq_conc = inverse_4pl(lloq_sig, *params)
            if not (np.isfinite(lloq_conc) and lloq_conc > 0):
                lloq_conc = None
        except (ValueError, ZeroDivisionError, OverflowError):
            lloq_conc = None

    # Generate smooth fitted curve
    conc_min = np.min(std_concs) * 0.3
    conc_max = np.max(std_concs) * 3
    x_smooth = np.logspace(np.log10(conc_min), np.log10(conc_max), 200)
    y_smooth = four_pl(x_smooth, *params)

    # ── Plot ──────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(8, 5.5))

    # Detection range shading
    if lloq_sig is not None and uloq_sig is not None:
        ax.axhspan(lloq_sig, uloq_sig, alpha=0.08, color='#2244AA', zorder=0)
        # "In Detection Range" label at left edge, vertically centered
        mid_sig = np.sqrt(lloq_sig * uloq_sig)  # geometric mean for log scale
        ax.text(conc_min * 0.35, mid_sig, 'In Detection Range',
                fontsize=7.5, color='#1a3a8a', style='italic', ha='left', va='center')

    # ULOQ line (top)
    if uloq_sig is not None:
        ax.axhline(y=uloq_sig, color='#1a3a8a', linestyle=':', linewidth=1.2, zorder=1)
        ax.text(conc_min * 0.35, uloq_sig * 1.25, 'Above Detection Range',
                fontsize=7.5, color='#1a3a8a', style='italic', ha='left', va='bottom')

    # LLOQ line (bottom)
    if lloq_sig is not None:
        ax.axhline(y=lloq_sig, color='#1a3a8a', linestyle=':', linewidth=1.2, zorder=1)
        ax.text(conc_max * 1.5, lloq_sig * 0.80, 'Below Detection Range',
                fontsize=7.5, color='#1a3a8a', style='italic', ha='right', va='top')

    # Fitted curve
    ax.plot(x_smooth, y_smooth, '-', color='#1a3a8a', linewidth=1.5, zorder=3)

    # Observed data points
    ax.scatter(std_concs, std_sigs, s=35, color='#1a3a8a', zorder=4,
               edgecolors='#0a2060', linewidths=0.5, label='Standards')

    # Log-log scale
    ax.set_xscale('log')
    ax.set_yscale('log')

    # Axis limits — leave room for labels
    ax.set_xlim(conc_min * 0.25, conc_max * 4)
    all_sigs = list(std_sigs)
    if lloq_sig is not None:
        all_sigs.append(lloq_sig)
    if uloq_sig is not None:
        all_sigs.append(uloq_sig)
    sig_min = min(s for s in all_sigs if s > 0) * 0.4
    sig_max = max(all_sigs) * 3
    ax.set_ylim(sig_min, sig_max)

    # Axis formatting
    ax.set_xlabel('Concentration', fontsize=10, fontweight='bold')
    ax.set_ylabel('Signal', fontsize=10, fontweight='bold')

    title_str = f"Plate {plate}, Spot {spot}"
    if group:
        title_str += f" — {group}"
    ax.set_title(title_str, fontsize=11, fontweight='bold', pad=10)

    # Tick formatting — show 10^n style
    for axis in [ax.xaxis, ax.yaxis]:
        axis.set_major_formatter(ticker.LogFormatterSciNotation())

    ax.tick_params(which='both', direction='in', top=True, right=True)
    ax.grid(True, which='major', alpha=0.15, linewidth=0.5)

    # Info box with Calc. Low / High
    info_lines = []
    if lloq_conc is not None:
        info_lines.append(f"Calc. Low    {lloq_conc:.2f}")
    info_lines.append(f"Calc. High   {uloq_conc:.0f}")
    if info_lines:
        box_text = '\n'.join(info_lines)
        props = dict(boxstyle='round,pad=0.4', facecolor='white', edgecolor='#666666', alpha=0.9)
        ax.text(0.98, 0.98, box_text, transform=ax.transAxes, fontsize=8,
                verticalalignment='top', horizontalalignment='right',
                bbox=props, family='monospace')

    # Legend
    ax.legend(loc='lower right', fontsize=8, framealpha=0.9)

    plt.tight_layout()

    # Save — sanitize name components so characters like '/' don't split the path
    def _safe(s):
        return re.sub(r'[^\w\-.]', '_', str(s))
    fname = f"chart_P{_safe(plate)}_S{_safe(spot)}{'_' + _safe(group) if group else ''}.png"
    fpath = os.path.join(tmp_dir, fname)
    fig.savefig(fpath, dpi=96, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return fpath


def generate_overlay_chart(results, tmp_dir, qc_overlay_points=None, qc_expected_concentrations=None):
    """
    Generate an overlay plot showing all fitted standard curves on one chart.
    Each spot/group gets its own color. QC points (if provided) are overlaid
    as star markers using their corrected concentration and original signal.
    Expected QC concentrations (if provided) are shown as ±30% vertical bands.
    Returns path to saved PNG.
    """
    fitted = [r for r in results if r.get('params') is not None]
    if not fitted:
        return None

    fig, ax = plt.subplots(figsize=(12, 7.5))

    cmap = plt.cm.get_cmap('tab10')

    # Build stable group→color map (first-seen order) so bands match curves
    _grp_color_map = {}
    _ci = 0
    for res in fitted:
        g = res.get('group', '') or ''
        if g not in _grp_color_map:
            _grp_color_map[g] = cmap(_ci % 10)
            _ci += 1
    colors = [_grp_color_map.get(res.get('group', '') or '', cmap(i % 10))
              for i, res in enumerate(fitted)]

    global_conc_min = np.inf
    global_conc_max = 0

    for idx, res in enumerate(fitted):
        params = res['params']
        standards = res.get('standards', [])
        std_concs = np.array([s['conc'] for s in standards if s['conc'] > 0])
        std_sigs = np.array([s['signal'] for s in standards if s['conc'] > 0])
        if len(std_concs) == 0:
            continue

        color = colors[idx]
        spot = res['spot']
        group = res.get('group', '')
        plate = res['plate']
        label = f"Spot {spot}"
        if group:
            label += f" ({group})"
        if len(set(r['plate'] for r in fitted)) > 1:
            label = f"P{plate} " + label

        cmin, cmax = np.min(std_concs), np.max(std_concs)
        global_conc_min = min(global_conc_min, cmin)
        global_conc_max = max(global_conc_max, cmax)

        # Smooth fitted curve
        x_smooth = np.logspace(np.log10(cmin * 0.3), np.log10(cmax * 3), 200)
        y_smooth = four_pl(x_smooth, *params)
        ax.plot(x_smooth, y_smooth, '-', color=color, linewidth=1.5, label=label, zorder=3)

        # Observed points
        ax.scatter(std_concs, std_sigs, s=25, color=color, zorder=4,
                   edgecolors='black', linewidths=0.3, alpha=0.8)

    # Shared QC level color palette
    qc_cmap = plt.cm.get_cmap('Set1')
    qc_level_colors = {level: qc_cmap(i % 9) for i, level in enumerate(QC_LEVELS)}

    # Per-group ±30% expected concentration bands, color-matched to each group's curve
    _exp_dict = qc_expected_concentrations if isinstance(qc_expected_concentrations, dict) else (
        {} if qc_expected_concentrations is None else {'': qc_expected_concentrations})
    _n_qc_bands = 0
    for _grp, _exp_conc in _exp_dict.items():
        if _exp_conc is None or not np.isfinite(float(_exp_conc)) or float(_exp_conc) <= 0:
            continue
        _exp_conc = float(_exp_conc)
        _band_clr = _grp_color_map.get(_grp, 'steelblue')
        _lo, _hi = _exp_conc * 0.70, _exp_conc * 1.30
        _lbl = f"{_grp} ±30% ({_exp_conc:.3g})" if _grp and _grp != '_default' else f"QC ±30% ({_exp_conc:.3g})"
        ax.axvspan(_lo, _hi, alpha=0.15, color=_band_clr, zorder=1, label=_lbl)
        ax.axvline(_exp_conc, color=_band_clr, linewidth=1.0, linestyle='--', zorder=2)
        global_conc_min = min(global_conc_min, _lo)
        global_conc_max = max(global_conc_max, _hi)
        _n_qc_bands += 1

    # QC overlay points (corrected conc vs original signal)
    if qc_overlay_points:
        plotted_levels = set()
        for pt in qc_overlay_points:
            conc = pt.get('corrected_conc')
            sig = pt.get('signal')
            level = pt.get('level', pt.get('sample_name', ''))
            if conc is None or sig is None:
                continue
            if not (np.isfinite(conc) and np.isfinite(sig) and conc > 0 and sig > 0):
                continue
            qc_color = qc_level_colors.get(level, 'black')
            lbl = f"QC: {level}" if level not in plotted_levels else None
            ax.scatter(conc, sig, s=120, marker='*', color=qc_color, zorder=6,
                       edgecolors='black', linewidths=0.5, label=lbl, alpha=0.95)
            plotted_levels.add(level)
            # Expand axis range to include QC points
            global_conc_min = min(global_conc_min, conc)
            global_conc_max = max(global_conc_max, conc)

    ax.set_xscale('log')
    ax.set_yscale('log')

    ax.set_xlim(global_conc_min * 0.2, global_conc_max * 5)
    ax.set_xlabel('Concentration', fontsize=11, fontweight='bold')
    ax.set_ylabel('Signal', fontsize=11, fontweight='bold')
    ax.set_title('All Standard Curves — Overlay', fontsize=13, fontweight='bold', pad=12)

    for axis in [ax.xaxis, ax.yaxis]:
        axis.set_major_formatter(ticker.LogFormatterSciNotation())

    ax.tick_params(which='both', direction='in', top=True, right=True)
    ax.grid(True, which='major', alpha=0.15, linewidth=0.5)

    n_qc_pts  = len(set(pt['level'] for pt in (qc_overlay_points or []))) if qc_overlay_points else 0
    n_series = len(fitted) + _n_qc_bands + n_qc_pts
    if n_series <= 6:
        ax.legend(loc='lower right', fontsize=8, framealpha=0.9)
    else:
        ax.legend(loc='center left', bbox_to_anchor=(1.02, 0.5), fontsize=7.5,
                  framealpha=0.9, ncol=1 + n_series // 15)
        fig.subplots_adjust(right=0.78)

    plt.tight_layout()

    fpath = os.path.join(tmp_dir, 'overlay_all_curves.png')
    fig.savefig(fpath, dpi=96, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return fpath




def parse_msd_file(filepath):
    """
    Parse MSD .txt file. Handles 1, 4, or 10 spots per well and multi-plate.
    Returns: list of dicts [{plate_num, spots_per_well, data: {well_id: [signal_per_spot]}}]
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    plate_sections = re.split(r'(?=Plate\s*#\s*:)', content)
    if len(plate_sections) == 1 and 'Plate #' not in plate_sections[0][:200]:
        plate_sections = [content]

    plates = []
    for section in plate_sections:
        if '==========Data==' not in section:
            continue

        m_plate = re.search(r'Plate\s*#\s*:\s*(\d+)', section)
        plate_num = int(m_plate.group(1)) if m_plate else len(plates) + 1

        m_spots = re.search(r'Spots Per Well\s*:\s*(\d+)', section)
        n_spots = int(m_spots.group(1)) if m_spots else 1

        data_start = section.index('==========Data==')
        rest = section[data_start:]
        lines = rest.split('\n')

        data_lines = []
        started = False
        for line in lines:
            if '==========Data==' in line:
                started = True
                continue
            if started and '==========' in line:
                break
            if started:
                data_lines.append(line)

        well_data = {}
        current_row = None

        for line in data_lines:
            raw = line.rstrip()
            if not raw.strip():
                current_row = None
                continue

            parts = raw.split('\t')
            label = parts[0].strip()

            if label and len(label) == 1 and label.isalpha():
                current_row = label.upper()

            if current_row is None:
                continue

            vals = []
            for p in parts[1:]:
                p = p.strip()
                if p:
                    try:
                        vals.append(float(p))
                    except ValueError:
                        pass

            if not vals:
                continue

            for ci, v in enumerate(vals):
                well = f"{current_row}{ci + 1}"
                if well not in well_data:
                    well_data[well] = []
                well_data[well].append(v)

        plates.append({
            'plate_num': plate_num,
            'spots_per_well': n_spots,
            'data': well_data
        })

    return plates


# ═══════════════════════════════════════════════════════════════════════════════
# PLATE MAP PARSER (GRID FORMAT)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_plate_map_grid(filepath):
    """
    Parse grid-format plate map CSV. Supports multiple plates stacked
    vertically, separated by blank rows.

    Single plate:
        ,1,2,3,...,12
        A,800000,800000,fCtx,...
        B,200000,200000,fCtx,...
        ...H,...

    Multiple plates (blank row separates each):
        ,1,2,3,...,12
        A,800000,800000,fCtx,...
        ...H,...
                                    ← blank row
        ,1,2,3,...,12               ← optional repeated header
        A,500000,500000,Sample,...
        ...H,...

    Returns: dict {plate_number: [entries]}
      where each entry = {well, sample_type, concentration, sample_name}
    """
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        raw_lines = f.readlines()

    # Split into plate blocks on blank lines
    blocks = []
    raw_blocks = {}
    current_block = []
    for line in raw_lines:
        stripped = line.strip().replace(',', '').strip()
        if stripped == '':
            if current_block:
                blocks.append(current_block)
                current_block = []
        else:
            current_block.append(line)
    if current_block:
        blocks.append(current_block)

    # Parse each block as a plate grid
    all_plates = {}
    for plate_idx, block_lines in enumerate(blocks):
        plate_num = plate_idx + 1
        text = ''.join(block_lines)

        try:
            df = pd.read_csv(StringIO(text), index_col=0, dtype=str,
                             on_bad_lines='skip', sep=',', skipinitialspace=True)
        except Exception:
            # Fallback: strip trailing commas and retry
            cleaned = '\n'.join(l.rstrip().rstrip(',') for l in block_lines)
            try:
                df = pd.read_csv(StringIO(cleaned), index_col=0, dtype=str)
            except Exception:
                continue

        df.index = df.index.astype(str).str.strip().str.upper()
        df.columns = [str(c).strip() for c in df.columns]

        # Skip blocks that don't look like plate grids (need row letters A-H/A-P)
        valid_rows = [r for r in df.index if re.match(r'^[A-P]$', r)]
        if not valid_rows:
            continue

        entries = []
        for row_letter in valid_rows:
            for col_str in df.columns:
                raw = df.loc[row_letter, col_str]
                val = str(raw).strip() if pd.notna(raw) else ''
                well = f"{row_letter}{col_str}"

                if val == '' or val.lower() == 'nan':
                    entries.append({'well': well, 'sample_type': 'Empty',
                                    'concentration': np.nan, 'sample_name': '',
                                    'group': '_default'})
                    continue

                # Extract group prefix if present (e.g. "A:800000" → group="A", val="800000")
                # Supports multi-group for standards: "HTT1&HTT2:800000" → groups ["HTT1","HTT2"]
                groups_for_well = ['_default']
                if ':' in val:
                    parts = val.split(':', 1)
                    candidate_group = parts[0].strip()
                    candidate_val = parts[1].strip()
                    # Accept prefix if it looks like a short tag (not a full path or URL)
                    if len(candidate_group) <= 40 and candidate_val:
                        # Split on & to allow shared standards across multiple groups
                        sub_groups = [g.strip() for g in candidate_group.split('&') if g.strip()]
                        if sub_groups:
                            groups_for_well = sub_groups
                        val = candidate_val

                for group in groups_for_well:
                    if val.lower() in ('buffer only', 'blank', 'buffer', 'bg', 'background', '0'):
                        entries.append({'well': well, 'sample_type': 'Blank',
                                        'concentration': 0, 'sample_name': val,
                                        'group': group})
                        continue

                    try:
                        conc = float(val.replace(',', ''))
                        entries.append({'well': well, 'sample_type': 'Standard',
                                        'concentration': conc, 'sample_name': f'STD ({conc})',
                                        'group': group})
                        continue
                    except ValueError:
                        pass

                    entries.append({'well': well, 'sample_type': 'Unknown',
                                    'concentration': np.nan, 'sample_name': val,
                                    'group': group})

        all_plates[plate_num] = entries
        raw_blocks[plate_num] = block_lines
        groups = set(e['group'] for e in entries if e['group'] != '_default')
        group_str = f" | Groups: {', '.join(sorted(groups))}" if groups else ""
        print(f"  Plate {plate_num}: {sum(1 for e in entries if e['sample_type']=='Standard')} stds, "
              f"{sum(1 for e in entries if e['sample_type']=='Unknown')} unknowns, "
              f"{sum(1 for e in entries if e['sample_type']=='Blank')} blanks{group_str}")

    return all_plates, raw_blocks


def normalize_well(w):
    w = str(w).strip().upper()
    m = re.match(r'^([A-P])0*(\d+)$', w)
    return f"{m.group(1)}{int(m.group(2))}" if m else w


def parse_plate_dilution_factors(raw_value, n_plates):
    if raw_value is None:
        return {}
    if isinstance(raw_value, dict):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return {i + 1: float(raw_value) for i in range(n_plates)}

    if isinstance(raw_value, (list, tuple)):
        parts = [str(p).strip() for p in raw_value if str(p).strip() != '']
    else:
        parts = [p.strip() for p in str(raw_value).split(',') if p.strip() != '']

    if not parts:
        return {}

    try:
        values = [float(p) for p in parts]
    except ValueError:
        raise ValueError("Invalid dilution factors. Use numbers separated by commas.")

    if len(values) == 1:
        return {i + 1: values[0] for i in range(n_plates)}
    if len(values) == n_plates:
        return {i + 1: v for i, v in enumerate(values)}

    raise ValueError(f"Expected 1 or {n_plates} dilution factor(s), got {len(values)}.")


def calculate_lloq_signal(signals, lloq_method='current'):
    if not signals:
        return None
    values = [s for s in signals if np.isfinite(s)]
    if not values:
        return None
    mean_sig = np.mean(values)
    if lloq_method == '3xblank':
        return mean_sig * 3
    if len(values) > 1:
        return mean_sig + 10 * np.std(values, ddof=1)
    print("  ⚠ LLOQ not computed: only one finite blank replicate available "
          "(mean + 10×SD requires ≥2). Use '3×Blank Mean' method for single blanks.")
    return None


def parse_total_protein_csv(filepath):
    """
    Parse a total protein CSV file. Expects columns:
      'External Animal Number', 'Tissue Type', 'Total Protein Result'
    Multiple rows with the same (animal, tissue) are kept in order of appearance.
    Returns dict {(animal_str, tissue_str): [val1, val2, ...]}
    """
    df = pd.read_csv(filepath, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    required = {'External Animal Number', 'Tissue Type', 'Total Protein Result'}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Total protein CSV missing columns: {missing}")
    tp_map = {}
    for _, row in df.iterrows():
        animal = str(row['External Animal Number']).strip()
        tissue = str(row['Tissue Type']).strip()
        try:
            val = float(row['Total Protein Result'])
        except (ValueError, TypeError):
            continue
        key = (animal, tissue)
        if key not in tp_map:
            tp_map[key] = []
        tp_map[key].append(val)
    return tp_map


def _extract_animal_tissue(sample_name):
    """
    Flexible extraction of animal number and tissue from a sample name.
    - Strips any trailing _suffix (e.g. _P1, _rep2) before parsing
    - Splits by '-' and scans segments regardless of order or extras:
        Animal → first purely numeric segment        (e.g. '1001')
        Tissue → longest purely alphabetic segment   (e.g. 'fCtx' beats 'XX')
    - Returns (None, None) if no animal number found (e.g. HQC, Buffer Only)
    Handles any ordering or extra segments:
        fCtx-1001_P1, 1001-fCtx, fCtx-XX-1001, 1001-XX-fCtx, etc.
    """
    base = sample_name.strip().split('_')[0]   # drop _P1, _rep2, etc.
    segments = base.split('-')
    animal = next((s for s in segments if s.isdigit()), None)
    if animal is None:
        return None, None   # no animal ID → QC or non-sample name
    alpha_segs = [s for s in segments if s.isalpha()]
    tissue = max(alpha_segs, key=len) if alpha_segs else None
    return animal, tissue


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════



def _style_row(ws, row, max_col, fill=None, font=None):
    if font is None:
        font = DATA_FONT
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if fill:
            cell.fill = fill

def _header_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def _section_title(ws, row, title, span=5):
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _resolve_dilution_factor(sample_name, group, plate,
                             qc_dilution_factors=None,
                             group_dilution_factors=None,
                             plate_dilution_factors=None):
    """Return the dilution factor for a sample using priority: QC > group > plate.

    Returns the resolved factor (float, default 1.0).
    """
    plate_dilution_factors = plate_dilution_factors or {}
    grp_key = group if group and group != '_default' else ''

    # QC level takes highest priority when a matching factor exists
    if qc_dilution_factors:
        level = _identify_qc_level(sample_name)
        if level:
            grp_qc = qc_dilution_factors.get(grp_key, {})
            if level in grp_qc:
                return grp_qc[level]

    # Group-level factor next
    if grp_key and group_dilution_factors and grp_key in group_dilution_factors:
        return group_dilution_factors[grp_key]

    # Fall back to per-plate factor
    return plate_dilution_factors.get(plate, 1.0)


def _compute_qc_summary(results, qc_dilution_factors, qc_expected_concentrations):
    """Aggregate QC replicate wells and compute corrected concentrations and recovery.

    Returns (qc_summary_rows, qc_overlay_points) — both lists of dicts.
    Called by both create_output (Excel) and generate_html_report (HTML) to
    avoid duplicating this logic.
    """
    qc_summary_rows = []
    qc_overlay_points = []
    if not qc_dilution_factors:
        return qc_summary_rows, qc_overlay_points

    qc_groups = defaultdict(list)
    for res in results:
        grp = res.get('group', '') or ''
        group_qc = qc_dilution_factors.get(grp, {})
        for unk in res.get('unknowns', []):
            sname = unk.get('sample_name', '')
            level = _identify_qc_level(sname)
            if level and level in group_qc:
                key = (sname, grp, res['plate'])
                qc_groups[key].append({
                    'signal': unk['signal'],
                    'interp_conc': unk['interp_conc'],
                    'level': level,
                })

    for (sname, grp, plate), entries in sorted(qc_groups.items()):
        sigs  = [e['signal']       for e in entries if np.isfinite(e['signal'])]
        concs = [e['interp_conc']  for e in entries if np.isfinite(e['interp_conc'])]
        level = entries[0]['level']
        qc_factor = qc_dilution_factors.get(grp, {}).get(level, 1.0)
        avg_sig  = np.mean(sigs)  if sigs  else np.nan
        avg_conc = np.mean(concs) if concs else np.nan
        corrected = avg_conc * qc_factor if np.isfinite(avg_conc) else np.nan
        exp_conc = (
            (qc_expected_concentrations or {}).get(grp)
            if isinstance(qc_expected_concentrations, dict)
            else qc_expected_concentrations
        )
        recovery = (corrected / exp_conc * 100
                    if exp_conc and np.isfinite(corrected) else np.nan)
        row = {
            'sample_name': sname, 'level': level, 'plate': plate, 'group': grp,
            'avg_signal': avg_sig, 'corrected_conc': corrected, 'recovery': recovery,
        }
        qc_summary_rows.append(row)
        if np.isfinite(corrected) and np.isfinite(avg_sig) and corrected > 0 and avg_sig > 0:
            qc_overlay_points.append({**row, 'signal': avg_sig})

    return qc_summary_rows, qc_overlay_points


def create_output(results, output_path, msd_path, raw_plate_blocks, units=None, cv_threshold=25, plate_dilution_factors=None, lloq_method='current', total_protein_map=None, qc_dilution_factors=None, qc_expected_concentrations=None, group_dilution_factors=None):
    wb = Workbook()
    wb.remove(wb.active)
    tmp_dir = tempfile.mkdtemp(prefix='msd_charts_')

    # Pre-collect QC overlay points (corrected conc + signal) for overlay chart
    qc_summary_rows, qc_overlay_points = _compute_qc_summary(
        results, qc_dilution_factors, qc_expected_concentrations)

    # Pre-generate all charts before Excel writing (sequential — matplotlib mathtext is not thread-safe)
    overlay_path = generate_overlay_chart(
        results, tmp_dir,
        qc_overlay_points if qc_overlay_points else None,
        qc_expected_concentrations if qc_expected_concentrations else None
    )
    chart_map = {id(res): generate_std_curve_chart(res, tmp_dir, lloq_method) for res in results}

    unit_suffix = f" ({units})" if units else ""
    interp_header = f"Interp. Conc.{unit_suffix}"
    avg_interp_header = f"Avg Interp. Conc.{unit_suffix}"
    corrected_header = f"Corrected Avg Interp. Conc.{unit_suffix}"
    cv_threshold = float(cv_threshold) if cv_threshold is not None else 25.0
    plate_dilution_factors = plate_dilution_factors or {}
    lloq_method = lloq_method or 'current'
    lloq_method_label = "3× Blank Mean" if lloq_method == '3xblank' else "Blank Mean + 10×SD"

    # ── Summary Sheet ─────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    # Row 1: LLOQ method metadata
    ws.cell(row=1, column=1, value="LLOQ Method:").font = SECTION_FONT
    ws.cell(row=1, column=2, value=lloq_method_label)
    headers = ["Plate", "Spot", "Group", "Min (a)", "Hill Slope (b)", "EC50 (c)", "Max (d)", "LLOQ Signal", "LLOQ Conc", "R²", "Flags", "Status"]
    _header_row(ws, 2, headers)

    for ri, res in enumerate(results, 3):
        vals = [res['plate'], res['spot'], res.get('group', '')]
        if res['params'] is not None:
            a, b, c, d = res['params']
            vals += [round(a, 2), round(b, 4), round(c, 4), round(d, 2)]
        else:
            vals += ["N/A"] * 4

        # Calculate LLOQ
        lloq_sig_val = "N/A"
        lloq_conc_val = "N/A"
        lloq_sig = res.get('lloq_sig')
        if lloq_sig is not None:
            lloq_sig_val = round(lloq_sig, 1)
            if res['params'] is not None:
                try:
                    lloq_conc = inverse_4pl(lloq_sig, *res['params'])
                    if np.isfinite(lloq_conc) and lloq_conc > 0:
                        lloq_conc_val = round(lloq_conc, 4)
                except (ValueError, ZeroDivisionError, OverflowError):
                    pass
        vals.append(lloq_sig_val)
        vals.append(lloq_conc_val)

        if res['params'] is not None:
            vals += [round(res['r2'], 6)]
            flag_text = "No standards" if res.get('no_standards') else ""
            vals.append(flag_text)
            r2v = res['r2']
            vals.append("Good" if r2v >= R2_GOOD else ("Acceptable" if r2v >= R2_ACCEPTABLE else ("Negative R²" if r2v < 0 else "Poor")))
        else:
            vals += ["N/A", "", "Failed"]

        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)
        status = ws.cell(row=ri, column=12)
        status.font = PASS_FONT if status.value == "Good" else (WARN_FONT if status.value == "Acceptable" else FAIL_FONT)
        _style_row(ws, ri, len(headers))

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ── QC Recovery Table on Summary sheet ───────────────────────────
    next_row = len(results) + 3
    if qc_summary_rows:
        _section_title(ws, next_row, "QC Recovery")
        next_row += 1
        qc_h = ["Sample Name", "Level", "Plate", "Group", "Avg Signal",
                corrected_header, "Expected Conc.", "% Recovery"]
        _header_row(ws, next_row, qc_h)
        next_row += 1
        for qr in qc_summary_rows:
            ws.cell(row=next_row, column=1, value=qr['sample_name'])
            ws.cell(row=next_row, column=2, value=qr['level'])
            ws.cell(row=next_row, column=3, value=qr['plate'])
            ws.cell(row=next_row, column=4, value=qr['group'] or "")
            sig_cell = ws.cell(row=next_row, column=5,
                               value=round(qr['avg_signal'], 1) if np.isfinite(qr['avg_signal']) else "N/A")
            sig_cell.number_format = '#,##0'
            corr_cell = ws.cell(row=next_row, column=6,
                                value=round(qr['corrected_conc'], 4) if np.isfinite(qr['corrected_conc']) else "N/A")
            corr_cell.number_format = '#,##0.0000'
            exp_conc_val = (qc_expected_concentrations or {}).get(qr['group']) if isinstance(qc_expected_concentrations, dict) else qc_expected_concentrations
            exp_cell = ws.cell(row=next_row, column=7, value=exp_conc_val if exp_conc_val else "")
            if exp_conc_val:
                exp_cell.number_format = '#,##0.0###'
            rec_cell = ws.cell(row=next_row, column=8)
            if np.isfinite(qr['recovery']):
                rec_cell.value = round(qr['recovery'], 1)
                rec_cell.number_format = '0.0'
                rec_cell.font = PASS_FONT if QC_RECOVERY_LOW <= qr['recovery'] <= QC_RECOVERY_HIGH else FAIL_FONT
            else:
                rec_cell.value = "N/A"
            _style_row(ws, next_row, len(qc_h))
            next_row += 1
        for ci, w in enumerate([18, 10, 8, 10, 14, 24, 16, 12], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        next_row += 1  # blank row before chart

    # Overlay chart of all curves on Summary sheet (pre-generated above)
    if overlay_path:
        overlay_row = next_row + 1
        img = XlImage(overlay_path)
        img.width = 900
        img.height = 560
        ws.add_image(img, f"A{overlay_row}")

    # ── Per-Spot Detail Sheets ────────────────────────────────────────
    added_plates = set()
    for res in results:
        spot, plate = res['spot'], res['plate']
        group = res.get('group', '')
        g_suffix = f"_{group}" if group else ""
        sname = f"P{plate}_S{spot}{g_suffix}"[:31]
        ws = wb.create_sheet(sname)
        row = 1

        title_str = f"4PL Curve Fit — Plate {plate}, Spot {spot}"
        if group:
            title_str += f", Group {group}"
        _section_title(ws, row, title_str)
        row += 1
        param_names = ["Min (a)", "Hill Slope (b)", "EC50 (c)", "Max (d)", "R²"]
        for i, pname in enumerate(param_names):
            ws.cell(row=row, column=1, value=pname).font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            if res['params'] is not None:
                val = res['params'][i] if i < 4 else res['r2']
                ws.cell(row=row, column=2, value=round(val, 6))
            else:
                ws.cell(row=row, column=2, value="N/A")
            ws.cell(row=row, column=2).font = DATA_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = '0.000000'
            row += 1

        # Standards table (grouped by concentration = mean of replicates)
        row += 1
        _section_title(ws, row, "Standard Curve Data")
        row += 1
        _header_row(ws, row, ["Well(s)", "Concentration", "Mean Signal", "Fitted Signal", "% Recovery"])
        row += 1

        std_groups = {}
        for s in res.get('standards', []):
            key = s['conc']
            if key not in std_groups:
                std_groups[key] = {'wells': [], 'signals': [], 'conc': key}
            std_groups[key]['wells'].append(s['well'])
            std_groups[key]['signals'].append(s['signal'])

        for sg in sorted(std_groups.values(), key=lambda x: x['conc'], reverse=True):
            mean_sig = np.mean(sg['signals'])
            fitted = four_pl(sg['conc'], *res['params']) if res['params'] is not None else None
            recovery = (fitted / mean_sig * 100) if fitted and mean_sig != 0 else None

            ws.cell(row=row, column=1, value=', '.join(sg['wells']))
            ws.cell(row=row, column=2, value=sg['conc'])
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=round(mean_sig, 1))
            ws.cell(row=row, column=3).number_format = '#,##0.0'
            ws.cell(row=row, column=4, value=round(fitted, 1) if fitted else "N/A")
            ws.cell(row=row, column=4).number_format = '#,##0.0'
            if recovery:
                ws.cell(row=row, column=5, value=round(recovery, 1))
                ws.cell(row=row, column=5).number_format = '0.0'
            _style_row(ws, row, 5, fill=STD_FILL)
            row += 1

        # Individual standard points data (kept in columns G-I for reference)
        ind_start = row + 1
        ws.cell(row=ind_start, column=7, value="Conc").font = BOLD_FONT
        ws.cell(row=ind_start, column=8, value="Signal").font = BOLD_FONT
        ws.cell(row=ind_start, column=9, value="Fitted").font = BOLD_FONT
        irow = ind_start + 1
        for s in sorted(res.get('standards', []), key=lambda x: x['conc']):
            if s['conc'] > 0 and s['signal'] > 0:
                ws.cell(row=irow, column=7, value=s['conc'])
                ws.cell(row=irow, column=8, value=s['signal'])
                if res['params'] is not None:
                    fitted_val = four_pl(s['conc'], *res['params'])
                    if fitted_val > 0:
                        ws.cell(row=irow, column=9, value=round(fitted_val, 1))
                irow += 1

        # Blanks
        if res.get('blanks'):
            row += 2
            _section_title(ws, row, "Blanks / Background", 3)
            row += 1
            _header_row(ws, row, ["Well", "Sample Name", "Signal"])
            row += 1
            for bl in res['blanks']:
                ws.cell(row=row, column=1, value=bl['well'])
                ws.cell(row=row, column=2, value=bl.get('sample_name', ''))
                ws.cell(row=row, column=3, value=bl['signal'])
                _style_row(ws, row, 3, fill=BLANK_FILL)
                row += 1

        # Unknowns
        row += 2
        _section_title(ws, row, "Interpolated Unknowns")
        row += 1
        _header_row(ws, row, ["Well", "Sample Name", "Signal", interp_header, "Flag"])
        row += 1

        std_concs = [s['conc'] for s in res.get('standards', []) if s['conc'] > 0]
        uloq = max(std_concs) if std_concs else None
        lloq = min(std_concs) if std_concs else None

        # Use pre-computed LLOQ signal for this spot
        lloq_sig = res.get('lloq_sig')

        for unk in res.get('unknowns', []):
            ws.cell(row=row, column=1, value=unk['well'])
            ws.cell(row=row, column=2, value=unk.get('sample_name', ''))
            ws.cell(row=row, column=3, value=unk['signal'])
            ws.cell(row=row, column=3).number_format = '#,##0'
            c_val = unk['interp_conc']
            if c_val is not None and np.isfinite(c_val):
                ws.cell(row=row, column=4, value=round(c_val, 4))
                ws.cell(row=row, column=4).number_format = '#,##0.0000'
                # Check signal against LLOQ signal threshold first
                if lloq_sig and unk['signal'] < lloq_sig:
                    ws.cell(row=row, column=5, value="< LLOQ")
                    ws.cell(row=row, column=5).font = WARN_FONT
                elif uloq and c_val > uloq:
                    ws.cell(row=row, column=5, value="> ULOQ")
                    ws.cell(row=row, column=5).font = WARN_FONT
                elif lloq and c_val < lloq:
                    ws.cell(row=row, column=5, value="< LLOQ")
                    ws.cell(row=row, column=5).font = WARN_FONT
                else:
                    ws.cell(row=row, column=5, value="In Range")
                    ws.cell(row=row, column=5).font = PASS_FONT
            else:
                ws.cell(row=row, column=4, value="N/A")
                ws.cell(row=row, column=5, value="Out of Range")
                ws.cell(row=row, column=5).font = FAIL_FONT
            _style_row(ws, row, 5, fill=UNK_FILL)
            row += 1

        # Chart — matplotlib image (pre-generated in parallel)
        chart_path = chart_map.get(id(res))
        if chart_path:
            row += 2
            img = XlImage(chart_path)
            img.width = 680
            img.height = 470
            ws.add_image(img, f"A{row}")

        # Plate Map — add to the first sheet for each plate
        if plate not in added_plates and plate in raw_plate_blocks:
            row += 20  # Leave space after chart
            _section_title(ws, row, f"Plate {plate} Map")
            row += 1
            block_lines = raw_plate_blocks[plate]
            for ri, line in enumerate(block_lines, 1):
                parts = line.strip().split(',')
                for ci, part in enumerate(parts, 1):
                    ws.cell(row=row + ri - 1, column=ci, value=part.strip())
            # Adjust row counter
            row += len(block_lines)
            added_plates.add(plate)

        for ci, w in enumerate([14, 18, 14, 16, 14, 2, 14, 14, 14], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ── All Unknowns Combined ─────────────────────────────────────────
    ws_all = wb.create_sheet("All Unknowns")
    all_h = ["Sample Name", "Animal", "Tissue", "Plate", "Wells", "Avg Signal", avg_interp_header,
             "%CV", "Flag", "Dilution Factor", corrected_header, "Total Protein", "Normalized Protein Concentration"]
    _header_row(ws_all, 1, all_h)
    arow = 2
    # Track how many TP values have been consumed per (animal, tissue) key
    tp_index = defaultdict(int)

    # Collect all standards for uloq/lloq
    all_std_concs = set()
    for res in results:
        for s in res.get('standards', []):
            if s['conc'] > 0:
                all_std_concs.add(s['conc'])
    uloq = max(all_std_concs) if all_std_concs else None
    lloq = min(all_std_concs) if all_std_concs else None

    # Use pre-computed LLOQ signals; take the mean across all spots as a global threshold
    cached_lloq_sigs = [r['lloq_sig'] for r in results if r.get('lloq_sig') is not None]
    all_lloq_sig = float(np.mean(cached_lloq_sigs)) if cached_lloq_sigs else None

    # Group unknowns by (sample_name, group, plate)
    unknown_groups = defaultdict(list)
    for res in results:
        curve_group = res.get('group', '')
        plate = res['plate']
        for unk in res.get('unknowns', []):
            sample_name = unk.get('sample_name', '')
            key = (sample_name, curve_group, plate)
            unknown_groups[key].append({
                'well': unk['well'],
                'signal': unk['signal'],
                'interp_conc': unk['interp_conc']
            })

    for (sample_name, curve_group, plate) in sorted(unknown_groups.keys()):
        if _identify_qc_level(sample_name):
            continue  # QC samples reported in Summary sheet QC Recovery table
        group = unknown_groups[(sample_name, curve_group, plate)]
        signals = [g['signal'] for g in group if np.isfinite(g['signal'])]
        concs = [g['interp_conc'] for g in group if np.isfinite(g['interp_conc'])]
        avg_signal = np.mean(signals) if signals else np.nan
        avg_conc = np.mean(concs) if concs else np.nan
        wells = ', '.join(sorted(g['well'] for g in group))

        # Determine dilution factor: QC > group > plate
        factor = _resolve_dilution_factor(
            sample_name, curve_group, plate,
            qc_dilution_factors, group_dilution_factors, plate_dilution_factors)
        is_qc_factor = bool(_identify_qc_level(sample_name) and qc_dilution_factors) or \
                       bool(curve_group and group_dilution_factors and
                            (curve_group if curve_group != '_default' else '') in group_dilution_factors) or \
                       (plate in plate_dilution_factors)

        corrected_conc = avg_conc * factor if np.isfinite(avg_conc) else np.nan

        flag = ""
        if np.isfinite(avg_signal) and all_lloq_sig and avg_signal < all_lloq_sig:
            flag = "< LLOQ"
        elif np.isfinite(avg_conc):
            if uloq and avg_conc > uloq:
                flag = "> ULOQ"
            elif lloq and avg_conc < lloq:
                flag = "< LLOQ"
            else:
                flag = "In Range"
        else:
            flag = "Out of Range"

        cv = np.nan
        if len(concs) > 1 and np.isfinite(avg_conc) and avg_conc != 0:
            cv = np.std(concs, ddof=1) / avg_conc * 100

        animal, tissue = _extract_animal_tissue(sample_name)
        ws_all.cell(row=arow, column=1, value=sample_name)
        ws_all.cell(row=arow, column=2, value=animal or "")
        ws_all.cell(row=arow, column=3, value=tissue or "")
        ws_all.cell(row=arow, column=4, value=plate)
        ws_all.cell(row=arow, column=5, value=wells)
        ws_all.cell(row=arow, column=6, value=round(avg_signal, 1) if np.isfinite(avg_signal) else "N/A")
        ws_all.cell(row=arow, column=6).number_format = '#,##0'
        ws_all.cell(row=arow, column=7, value=round(avg_conc, 4) if np.isfinite(avg_conc) else "N/A")
        ws_all.cell(row=arow, column=7).number_format = '#,##0.0000'
        # %CV (col 8)
        cv_cell = ws_all.cell(row=arow, column=8)
        cv_cell.value = round(cv, 1) if np.isfinite(cv) else "N/A"
        cv_cell.number_format = '0.0'
        if np.isfinite(cv):
            cv_cell.fill = CV_BAD_FILL if cv > cv_threshold else CV_GOOD_FILL
        # Flag (col 9)
        ws_all.cell(row=arow, column=9, value=flag)
        cell_flag = ws_all.cell(row=arow, column=9)
        cell_flag.font = PASS_FONT if flag == "In Range" else (WARN_FONT if flag in ["> ULOQ", "< LLOQ"] else FAIL_FONT)
        # Dilution Factor (col 10)
        df_cell = ws_all.cell(row=arow, column=10)
        has_factor = is_qc_factor or (plate in plate_dilution_factors)
        df_cell.value = factor if has_factor else ""
        if has_factor:
            df_cell.number_format = '0.###'
        # Corrected Avg Interp. Conc. (col 11)
        corrected_cell = ws_all.cell(row=arow, column=11)
        corrected_cell.value = round(corrected_conc, 4) if np.isfinite(corrected_conc) else "N/A"
        corrected_cell.number_format = '#,##0.0000'
        # Total Protein (col 12) — consume values in order of appearance per (animal, tissue)
        tp_val = None
        tp_cell = ws_all.cell(row=arow, column=12)
        if total_protein_map and animal:
            tp_key = (animal, tissue)
            tp_list = total_protein_map.get(tp_key, [])
            idx = tp_index[tp_key]
            if idx < len(tp_list):
                tp_val = tp_list[idx]
                tp_index[tp_key] += 1
                tp_cell.value = round(tp_val, 4)
                tp_cell.number_format = '0.0000'
        # Normalized Protein Concentration (col 13)
        norm_cell = ws_all.cell(row=arow, column=13)
        if tp_val is not None and np.isfinite(corrected_conc) and tp_val != 0:
            norm_cell.value = round(corrected_conc / tp_val, 6)
            norm_cell.number_format = '0.000000'
        _style_row(ws_all, arow, len(all_h))
        arow += 1

    for ci in range(1, len(all_h) + 1):
        ws_all.column_dimensions[get_column_letter(ci)].width = 20

    # ── MSD Data Sheet ──────────────────────────────────────────────────────
    ws_msd = wb.create_sheet("MSD Data")
    with open(msd_path, 'r', encoding='utf-8', errors='replace') as f:
        msd_content = f.read()
    ws_msd.cell(row=1, column=1, value=msd_content)
    ws_msd.cell(row=1, column=1).alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
    ws_msd.column_dimensions['A'].width = 100

    wb.save(output_path)
    print(f"Saved: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def _open_file(path):
    """Open a file with the system default application (cross-platform)."""
    try:
        if platform.system() == 'Darwin':
            subprocess.Popen(['open', path])
        elif platform.system() == 'Windows':
            os.startfile(path)
        else:
            subprocess.Popen(['xdg-open', path])
    except Exception as e:
        print(f"Note: could not auto-open file: {e}")

def generate_html_report(results, html_path, msd_path, units=None,
                          qc_dilution_factors=None, qc_expected_concentrations=None,
                          plate_dilution_factors=None, lloq_method='current',
                          total_protein_map=None, excel_path=None,
                          group_dilution_factors=None, cv_threshold=25):
    """Generate a self-contained interactive HTML report alongside the Excel output."""
    try:
        import plotly.graph_objects as go
        import plotly.offline as poff
    except ImportError:
        print("Note: plotly not installed — HTML report skipped. Install with: pip install plotly")
        return

    plate_dilution_factors = plate_dilution_factors or {}
    unit_suffix = f" ({units})" if units else ""
    lloq_method_label = "3× Blank Mean" if lloq_method == '3xblank' else "Blank Mean + 10×SD"

    # ── QC summary rows (shared helper avoids duplication with create_output) ──
    qc_summary_rows, qc_overlay_points = _compute_qc_summary(
        results, qc_dilution_factors, qc_expected_concentrations)

    # ── Per-spot standard curve figures (built in parallel) ──────────────────
    def _build_curve_div(res):
        plate, spot, group = res['plate'], res['spot'], res.get('group', '')
        label = f"Plate {plate}, Spot {spot}" + (f", Group {group}" if group else "")
        fig = go.Figure()

        all_concs_pos, all_sigs_pos = [], []

        if res.get('standards'):
            std_groups_local = {}
            for s in res['standards']:
                key = s['conc']
                if key not in std_groups_local:
                    std_groups_local[key] = {'conc': key, 'signals': []}
                std_groups_local[key]['signals'].append(s['signal'])
            std_concs = sorted(std_groups_local.keys())
            std_means = [np.mean(std_groups_local[c]['signals']) for c in std_concs]
            all_concs_pos = [c for c in std_concs if c > 0]
            all_sigs_pos  = [s for s in std_means if s > 0]

            # Individual replicate points (smaller, semi-transparent)
            rep_concs, rep_sigs = [], []
            for c in std_concs:
                for sig in std_groups_local[c]['signals']:
                    rep_concs.append(c)
                    rep_sigs.append(sig)
            fig.add_trace(go.Scatter(
                x=rep_concs, y=rep_sigs,
                mode='markers', name='Replicates',
                marker=dict(color='#2F5496', size=6, symbol='circle', opacity=0.4),
                hovertemplate='Conc: %{x:.4g}<br>Signal: %{y:,.0f}<extra>Replicate</extra>'
            ))

            # Means (larger, opaque)
            fig.add_trace(go.Scatter(
                x=std_concs, y=std_means,
                mode='markers', name='Std Means',
                marker=dict(color='#2F5496', size=9, symbol='circle'),
                hovertemplate='Conc: %{x:.4g}<br>Signal: %{y:,.0f}<extra>Std Mean</extra>'
            ))

        lloq_sig = res.get('lloq_sig')
        if res['params'] is not None:
            concs_for_fit = [s['conc'] for s in res.get('standards', []) if s['conc'] > 0]
            if concs_for_fit:
                c_min, c_max = min(concs_for_fit), max(concs_for_fit)
                x_fit = np.logspace(np.log10(c_min * 0.5), np.log10(c_max * 2), 80)
                y_fit = four_pl(x_fit, *res['params'])
                all_sigs_pos += [v for v in y_fit if v > 0]
                fig.add_trace(go.Scatter(
                    x=list(x_fit), y=list(y_fit),
                    mode='lines', name='4PL Fit',
                    line=dict(color='#E06C4A', width=2),
                    hovertemplate='Conc: %{x:.4g}<br>Signal: %{y:,.0f}<extra>4PL Fit</extra>'
                ))

            if lloq_sig is not None and lloq_sig > 0:
                fig.add_hline(y=lloq_sig, line=dict(color='#F4A522', dash='dash', width=1.5),
                              annotation_text=f'LLOQ signal: {lloq_sig:,.0f}',
                              annotation_position='bottom right')
                all_sigs_pos.append(lloq_sig)
                # Vertical line at the interpolated LLOQ concentration
                try:
                    lloq_conc = inverse_4pl(lloq_sig, *res['params'])
                    if np.isfinite(lloq_conc) and lloq_conc > 0:
                        fig.add_vline(x=lloq_conc,
                                      line=dict(color='#F4A522', dash='dot', width=1.5),
                                      annotation_text=f'LLOQ: {lloq_conc:.4g}',
                                      annotation_position='top right')
                        all_concs_pos.append(lloq_conc)
                except Exception:
                    pass

        # ── Unknown sample scatter points ─────────────────────────────────────
        _factor = _resolve_dilution_factor(
            '', group, plate,
            None, group_dilution_factors, plate_dilution_factors)
        _unk_xs, _unk_ys, _unk_names = [], [], []
        for u in res.get('unknowns', []):
            sname = u['sample_name']
            if _identify_qc_level(sname):
                continue
            sig = u['signal']
            conc = u.get('interp_conc', np.nan)
            if np.isfinite(sig) and sig > 0 and np.isfinite(conc) and conc > 0:
                _unk_xs.append(conc * _factor)
                _unk_ys.append(sig)
                _unk_names.append(sname)
        sample_trace_idx = len(fig.data)
        has_samples = bool(_unk_xs)
        if has_samples:
            fig.add_trace(go.Scatter(
                x=_unk_xs, y=_unk_ys,
                mode='markers', name='Samples',
                marker=dict(color='#27AE60', size=8, symbol='circle-open',
                            line=dict(width=2, color='#27AE60')),
                text=_unk_names,
                hovertemplate='%{text}<br>Conc: %{x:.4g}<br>Signal: %{y:,.0f}<extra>Sample</extra>'
            ))

        x_range = ([np.log10(min(all_concs_pos)) - 0.25, np.log10(max(all_concs_pos)) + 0.25]
                   if all_concs_pos else None)
        y_range = ([np.log10(min(all_sigs_pos)) - 0.15, np.log10(max(all_sigs_pos)) + 0.15]
                   if all_sigs_pos else None)

        r2_str = f"R² = {res['r2']:.6f}" if res.get('r2') is not None else "Fit Failed"
        fig.update_layout(
            title=dict(text=f"{label}<br><sup>{r2_str}</sup>", x=0.5, font=dict(size=13)),
            xaxis=dict(title=f'Concentration{unit_suffix}', type='log',
                       showgrid=True, gridcolor='#ddd',
                       exponentformat='power', showexponent='all',
                       range=x_range),
            yaxis=dict(title='Signal', type='log',
                       showgrid=True, gridcolor='#ddd',
                       exponentformat='power', showexponent='all',
                       range=y_range),
            plot_bgcolor='white', paper_bgcolor='white',
            legend=dict(orientation='v', x=1.02, y=1),
            margin=dict(l=70, r=130, t=75, b=55),
            autosize=True, height=400
        )
        div_id = f"curve_p{plate}_s{spot}_{group or 'default'}"
        chart_html = fig.to_html(full_html=False, include_plotlyjs=False,
                                  div_id=div_id, config={'responsive': True})
        if has_samples:
            btn = (
                f'<button class="curve-toggle-btn" data-active="1" '
                f'onclick="msdToggleCurveSamples(this,\'{div_id}\',{sample_trace_idx})">'
                f'\U0001f441 Samples</button>'
            )
            chart_html = btn + chart_html
        return (label, chart_html)

    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor() as _pool:
        curve_divs = list(_pool.map(_build_curve_div, results))

    # ── Overlay figure ────────────────────────────────────────────────────────
    import json as _json
    overlay_fig = go.Figure()
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
              '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
    _group_trace_indices = defaultdict(list)        # group → [trace indices] for toggle buttons
    _overlay_sample_indices_by_group = defaultdict(list)  # group → sample trace indices only

    # Build a stable group→color map (one color per unique group, first-seen order)
    # so that curve traces and expected-concentration bands share the same color.
    _group_color_map = {}
    _col_idx = 0
    for res in results:
        if res['params'] is None:
            continue
        g = res.get('group', '') or ''
        if g not in _group_color_map:
            _group_color_map[g] = colors[_col_idx % len(colors)]
            _col_idx += 1

    _overlay_x_vals = []   # accumulated during first pass; avoids a second iteration
    for i, res in enumerate(results):
        if res['params'] is None:
            continue
        concs_for_fit = [s['conc'] for s in res.get('standards', []) if s['conc'] > 0]
        if not concs_for_fit:
            continue
        c_min, c_max = min(concs_for_fit), max(concs_for_fit)
        _overlay_x_vals.extend([c_min * 0.5, c_max * 2.0])
        x_fit = np.logspace(np.log10(c_min * 0.5), np.log10(c_max * 2), 80)
        y_fit = four_pl(x_fit, *res['params'])
        x_fit = list(x_fit)
        y_fit = list(y_fit)
        plate, spot, group = res['plate'], res['spot'], res.get('group', '')
        trace_label = f"P{plate} S{spot}" + (f" {group}" if group else "")
        color = _group_color_map.get(group, colors[i % len(colors)])
        _group_trace_indices[group or ''].append(len(overlay_fig.data))
        overlay_fig.add_trace(go.Scatter(
            x=x_fit, y=y_fit,
            mode='lines', name=trace_label,
            legendgroup=trace_label,
            line=dict(color=color, width=1.5),
            hovertemplate=f'%{{x:.4g}} → %{{y:,.0f}}<extra>{trace_label}</extra>'
        ))

        # Standard replicate points and means for this curve
        if res.get('standards'):
            _std_grps = {}
            for s in res['standards']:
                _std_grps.setdefault(s['conc'], []).append(s['signal'])
            _rep_xs = [c for c, sigs in _std_grps.items() for _ in sigs]
            _rep_ys = [sig for sigs in _std_grps.values() for sig in sigs]
            _mean_xs = list(_std_grps.keys())
            _mean_ys = [float(np.mean(v)) for v in _std_grps.values()]
            # Individual replicates (small, semi-transparent)
            _group_trace_indices[group or ''].append(len(overlay_fig.data))
            overlay_fig.add_trace(go.Scatter(
                x=_rep_xs, y=_rep_ys,
                mode='markers', name=f'{trace_label} replicates',
                legendgroup=trace_label, showlegend=False,
                marker=dict(color=color, size=5, symbol='circle', opacity=0.35),
                hovertemplate=f'Conc: %{{x:.4g}}<br>Signal: %{{y:,.0f}}<extra>{trace_label} replicate</extra>'
            ))
            # Means (larger, opaque)
            _group_trace_indices[group or ''].append(len(overlay_fig.data))
            overlay_fig.add_trace(go.Scatter(
                x=_mean_xs, y=_mean_ys,
                mode='markers', name=f'{trace_label} means',
                legendgroup=trace_label, showlegend=False,
                marker=dict(color=color, size=8, symbol='circle'),
                hovertemplate=f'Conc: %{{x:.4g}}<br>Mean signal: %{{y:,.0f}}<extra>{trace_label} mean</extra>'
            ))

        # Sample (unknown) scatter points for this curve, color-matched
        _grp_key = group if group and group != '_default' else ''
        if _grp_key and group_dilution_factors and _grp_key in group_dilution_factors:
            _factor = group_dilution_factors[_grp_key]
        else:
            _factor = (plate_dilution_factors or {}).get(plate, 1.0)
        _unk_xs, _unk_ys, _unk_names = [], [], []
        for u in res.get('unknowns', []):
            sname = u['sample_name']
            if _identify_qc_level(sname):
                continue
            sig = u['signal']
            conc = u.get('interp_conc', np.nan)
            if np.isfinite(sig) and sig > 0 and np.isfinite(conc) and conc > 0:
                _unk_xs.append(conc * _factor)
                _unk_ys.append(sig)
                _unk_names.append(sname)
        if _unk_xs:
            _overlay_x_vals.extend(_unk_xs)
            _sample_tidx = len(overlay_fig.data)
            _group_trace_indices[group or ''].append(_sample_tidx)
            _overlay_sample_indices_by_group[group or ''].append(_sample_tidx)
            overlay_fig.add_trace(go.Scatter(
                x=_unk_xs, y=_unk_ys,
                mode='markers', name=f'{trace_label} samples',
                legendgroup=trace_label,
                showlegend=False,
                marker=dict(color=color, size=7, symbol='circle-open',
                            line=dict(width=1.5, color=color)),
                text=_unk_names,
                hovertemplate='%{text}<br>Conc: %{x:.4g}<br>Signal: %{y:,.0f}<extra>' + trace_label + '</extra>'
            ))

    if qc_overlay_points:
        qc_level_colors = {'ULOQ': '#e41a1c', 'HQC': '#ff7f00', 'MQC': '#4daf4a',
                           'LQC': '#377eb8', 'LLOQ': '#984ea3'}
        # Group by (group, level) so each group's QC stars can be toggled independently
        qc_by_grp_level = defaultdict(list)
        for qp in qc_overlay_points:
            _qgrp = qp.get('group', '') or ''
            qc_by_grp_level[(_qgrp, qp['level'])].append(qp)
        for (_qgrp, level), pts in sorted(qc_by_grp_level.items()):
            xs = [p['corrected_conc'] for p in pts]
            ys = [p['signal'] for p in pts]
            names = [f"{p['sample_name']} (P{p['plate']})" for p in pts]
            _qgrp_label = f'{_qgrp} ' if _qgrp and _qgrp != '_default' else ''
            _group_trace_indices[_qgrp].append(len(overlay_fig.data))
            overlay_fig.add_trace(go.Scatter(
                x=xs, y=ys,
                mode='markers', name=f'QC {_qgrp_label}{level}',
                marker=dict(color=qc_level_colors.get(level, 'black'), size=12, symbol='star'),
                customdata=names,
                hovertemplate='Conc: %{x:.4g}<br>Signal: %{y:,.0f}<br>%{customdata}<extra>QC ' + _qgrp_label + level + '</extra>'
            ))

    # One LLOQ line per group label — averaged across all plates/spots sharing that label
    _group_shape_indices = defaultdict(list)   # group key → [layout.shapes indices]
    _lloq_group_palette = ['#E07B00', '#C0392B', '#1A7ABF', '#27AE60', '#8E44AD',
                           '#2C3E50', '#D35400', '#16A085', '#7F8C8D', '#F39C12']
    _lloq_by_group = defaultdict(lambda: {'sigs': [], 'concs': []})
    for res in results:
        if res.get('lloq_sig') is None or res['lloq_sig'] <= 0:
            continue
        g = res.get('group') or '_ungrouped'
        _lloq_by_group[g]['sigs'].append(res['lloq_sig'])
        if res['params'] is not None:
            try:
                lc = inverse_4pl(res['lloq_sig'], *res['params'])
                if np.isfinite(lc) and lc > 0:
                    _lloq_by_group[g]['concs'].append(
                        lc * (plate_dilution_factors or {}).get(res['plate'], 1.0))
            except Exception:
                pass

    _overlay_all_sigs = []
    for gi, (g_label, d) in enumerate(sorted(_lloq_by_group.items(),
                                              key=lambda kv: -np.mean(kv[1]['sigs']) if kv[1]['sigs'] else 0)):
        if not d['sigs']:
            continue
        avg_sig = float(np.mean(d['sigs']))
        _overlay_all_sigs.append(avg_sig)
        clr = _lloq_group_palette[gi % len(_lloq_group_palette)]
        display_name = g_label if g_label != '_ungrouped' else ''
        prefix = f'LLOQ ({display_name})' if display_name else 'LLOQ'
        if d['concs']:
            avg_conc = float(np.mean(d['concs']))
            conc_str = f'{avg_conc:.4g}' + (f' {units}' if units else '')
            ann = f'{prefix}: {avg_sig:,.0f} (signal) | {conc_str} (conc)'
        else:
            ann = f'{prefix}: {avg_sig:,.0f} (signal)'
        # Dashed horizontal line — track layout shape index for toggle
        _curve_grp_key = '' if g_label == '_ungrouped' else g_label
        _lloq_shape_idx = len(overlay_fig.layout.shapes)
        overlay_fig.add_hline(
            y=avg_sig,
            line=dict(color=clr, dash='dash', width=2),
        )
        _group_shape_indices[_curve_grp_key].append(_lloq_shape_idx)
        # Dummy trace in legend2 — positioned near the LLOQ lines (bottom of chart)
        _group_trace_indices[_curve_grp_key].append(len(overlay_fig.data))
        overlay_fig.add_trace(go.Scatter(
            x=[None], y=[None],
            mode='lines',
            name=ann,
            line=dict(color=clr, dash='dash', width=2),
            showlegend=True,
            legend='legend2',
        ))

    # Per-group ±30% expected concentration bands, color-matched to each group's curve
    if isinstance(qc_expected_concentrations, dict):
        for _grp, _exp_conc in qc_expected_concentrations.items():
            if not _exp_conc or not np.isfinite(float(_exp_conc)) or float(_exp_conc) <= 0:
                continue
            _exp_conc = float(_exp_conc)
            _band_color = _group_color_map.get(_grp, 'steelblue')
            _lo, _hi = _exp_conc * 0.7, _exp_conc * 1.3
            _grp_label = f'{_grp} ' if _grp and _grp != '_default' else ''
            _vrect_shape_idx = len(overlay_fig.layout.shapes)
            overlay_fig.add_vrect(
                x0=_lo, x1=_hi,
                fillcolor=_band_color, opacity=0.15,
                layer='below', line_width=0,
                annotation_text=f'{_grp_label}±30% ({_lo:.4g}–{_hi:.4g})',
                annotation_position='top right',
            )
            _group_shape_indices[_grp].append(_vrect_shape_idx)
    elif qc_expected_concentrations and float(qc_expected_concentrations) > 0:
        # Legacy single-value fallback
        _exp_conc = float(qc_expected_concentrations)
        _lo, _hi = _exp_conc * 0.7, _exp_conc * 1.3
        overlay_fig.add_vrect(
            x0=_lo, x1=_hi,
            fillcolor='steelblue', opacity=0.15,
            layer='below', line_width=0,
            annotation_text=f'±30% ({_lo:,.1f}–{_hi:,.1f})',
            annotation_position='top right'
        )

    # Extend x-range accumulator with QC points (the curve/sample values were
    # already collected during the first trace-building loop above).
    if qc_overlay_points:
        for _qp in qc_overlay_points:
            if np.isfinite(_qp['corrected_conc']) and _qp['corrected_conc'] > 0:
                _overlay_x_vals.append(_qp['corrected_conc'])
    overlay_x_range = None
    if _overlay_x_vals:
        overlay_x_range = [np.log10(min(_overlay_x_vals)) - 0.2,
                           np.log10(max(_overlay_x_vals)) + 0.2]

    # Estimate paper-y coordinate for the LLOQ legend so it sits next to the lines.
    # Collect all visible signal values to infer the auto y-axis log range.
    _all_visible_sigs = []
    for res in results:
        for s in res.get('standards', []):
            if s.get('signal', 0) > 0:
                _all_visible_sigs.append(s['signal'])
        if res.get('lloq_sig') and res['lloq_sig'] > 0:
            _all_visible_sigs.append(res['lloq_sig'])
    if _all_visible_sigs and _overlay_all_sigs:
        _log_min = np.log10(min(_all_visible_sigs))
        _log_max = np.log10(max(_all_visible_sigs))
        _avg_lloq = float(np.mean(_overlay_all_sigs))
        _lloq_paper = ((np.log10(_avg_lloq) - _log_min) / (_log_max - _log_min)
                       if _log_max > _log_min else 0.15)
        _lloq_legend_y = float(np.clip(_lloq_paper, 0.05, 0.6))
    else:
        _lloq_legend_y = 0.15

    overlay_fig.update_layout(
        title=dict(text='Standard Curve Overlay', x=0.5),
        xaxis=dict(title=f'Concentration{unit_suffix}', type='log',
                   showgrid=True, gridcolor='#eee',
                   exponentformat='power', showexponent='all',
                   range=overlay_x_range),
        yaxis=dict(title='Signal', type='log',
                   showgrid=True, gridcolor='#eee',
                   exponentformat='power', showexponent='all'),
        plot_bgcolor='white', paper_bgcolor='white',
        # Main legend: horizontal, wraps into multiple rows above the plot
        legend=dict(
            orientation='h',
            x=0, xanchor='left',
            y=1.0, yanchor='bottom',
            itemclick='toggle', itemdoubleclick='toggleothers',
            font=dict(size=11),
            tracegroupgap=4,
        ),
        # LLOQ legend: pinned to bottom-right of the chart area
        legend2=dict(
            orientation='v',
            x=1.02, xanchor='left',
            y=0, yanchor='bottom',
            bgcolor='rgba(0,0,0,0)',
            borderwidth=0,
            font=dict(size=10),
        ),
        margin=dict(l=60, r=220, t=110, b=60),
        height=644
    )
    overlay_div = overlay_fig.to_html(full_html=False, include_plotlyjs=False,
                                       div_id='overlay_chart', config={'responsive': True})

    # ── Group toggle button bar (shown above the overlay chart) ──────────────
    _all_grp_indices = [idx for idxs in _group_trace_indices.values() for idx in idxs]
    _all_shape_indices = [idx for idxs in _group_shape_indices.values() for idx in idxs]
    _overlay_btns = ''
    if len(_group_trace_indices) > 1:
        _bs = ("padding:5px 14px;border:none;border-radius:4px;cursor:pointer;"
               "font-size:12px;font-weight:500;transition:opacity 0.15s;")
        _btn_parts = [
            '<div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px;align-items:center;">',
            '<span style="font-size:12px;color:#555;font-weight:600;margin-right:4px;">Groups:</span>',
            f'<button style="{_bs}background:#3a506b;color:white;" '
            f'onclick="msdOverlayAll(true)">Show All</button>',
            f'<button style="{_bs}background:#888;color:white;" '
            f'onclick="msdOverlayAll(false)">Hide All</button>',
        ]
        for _grp in sorted(_group_trace_indices.keys()):
            _tidxs = _group_trace_indices[_grp]
            _sidxs = _group_shape_indices.get(_grp, [])
            _display = _grp if _grp and _grp != '_default' else 'Default'
            _clr = _group_color_map.get(_grp, '#3a506b')
            _btn_parts.append(
                f'<button data-active="1" '
                f'style="{_bs}background:{_clr};color:white;" '
                f'onclick="msdToggleGrp(this,{_json.dumps(_tidxs)},{_json.dumps(_sidxs)})">'
                f'{_display}</button>'
            )
        _btn_parts.append('</div>')
        # Per-group Samples toggle row (only shown when at least one group has samples)
        if _overlay_sample_indices_by_group:
            _btn_parts.append(
                '<div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px;align-items:center;">'
                '<span style="font-size:12px;color:#555;font-weight:600;margin-right:4px;">Samples:</span>'
            )
            for _grp in sorted(_overlay_sample_indices_by_group.keys()):
                _sidxs = _overlay_sample_indices_by_group[_grp]
                _clr = _group_color_map.get(_grp, '#27AE60')
                _display = _grp if _grp and _grp != '_default' else 'Default'
                _btn_parts.append(
                    f'<button data-active="1" '
                    f'style="{_bs}background:{_clr};color:white;opacity:0.75;" '
                    f'onclick="msdToggleGrp(this,{_json.dumps(_sidxs)},[])">'
                    f'\U0001f441️ {_display}</button>'
                )
            _btn_parts.append('</div>')
        _overlay_btns = ''.join(_btn_parts)

    # ── Summary table rows ────────────────────────────────────────────────────
    summary_rows_html = []
    for res in results:
        plate, spot, group = res['plate'], res['spot'], res.get('group', '')
        a = b = c = d = r2 = lloq_sig_disp = lloq_conc_disp = status = flags = 'N/A'
        if res['params'] is not None:
            a, b, c, d = [f"{v:.4g}" for v in res['params']]
            r2 = f"{res['r2']:.6f}"
            r2_val = res['r2']
            status = ('Good' if r2_val >= R2_GOOD
                      else 'Acceptable' if r2_val >= R2_ACCEPTABLE
                      else 'Negative R²' if r2_val < 0
                      else 'Poor')
            flags = ''
        else:
            status = 'Failed'
            flags = 'No standards' if res.get('no_standards') else ''
        lloq_sig = res.get('lloq_sig')
        if lloq_sig is not None:
            lloq_sig_disp = f"{lloq_sig:,.1f}"
            if res['params'] is not None:
                try:
                    lconc = inverse_4pl(lloq_sig, *res['params'])
                    if np.isfinite(lconc) and lconc > 0:
                        lloq_conc_disp = f"{lconc:.4g}"
                except Exception:
                    pass
        status_class = {'Good': 'status-good', 'Acceptable': 'status-warn',
                        'Poor': 'status-fail', 'Negative R²': 'status-fail',
                        'Failed': 'status-fail'}.get(status, '')
        summary_rows_html.append(
            f"<tr><td>{plate}</td><td>{spot}</td><td>{group}</td>"
            f"<td>{a}</td><td>{b}</td><td>{c}</td><td>{d}</td>"
            f"<td>{lloq_sig_disp}</td><td>{lloq_conc_disp}</td><td>{r2}</td>"
            f"<td>{flags}</td><td class='{status_class}'>{status}</td></tr>"
        )

    # ── QC Recovery table HTML ────────────────────────────────────────────────
    qc_table_html = ''
    if qc_summary_rows:
        qc_hdr = f"Corrected Avg Interp. Conc.{unit_suffix}"
        qc_rows_html = []
        for qr in qc_summary_rows:
            rec = qr['recovery']
            if np.isfinite(rec):
                rec_class = 'status-good' if 70 <= rec <= 130 else 'status-fail'
                rec_str = f"{rec:.1f}%"
            else:
                rec_class = ''
                rec_str = 'N/A'
            avg_sig_str = f"{qr['avg_signal']:,.1f}" if np.isfinite(qr['avg_signal']) else 'N/A'
            corr_str = f"{qr['corrected_conc']:.4g}" if np.isfinite(qr['corrected_conc']) else 'N/A'
            exp_str_val = (qc_expected_concentrations or {}).get(qr['group']) if isinstance(qc_expected_concentrations, dict) else qc_expected_concentrations
            exp_str = f"{exp_str_val:.4g}" if exp_str_val else ''
            qc_rows_html.append(
                f"<tr><td>{qr['sample_name']}</td><td>{qr['level']}</td>"
                f"<td>{qr['plate']}</td><td>{qr['group'] or ''}</td>"
                f"<td>{avg_sig_str}</td><td>{corr_str}</td><td>{exp_str}</td>"
                f"<td class='{rec_class}'>{rec_str}</td></tr>"
            )
        qc_table_html = f"""
    <h2>QC Recovery</h2>
    <div class="table-wrap">
    <table id="qcTable" class="data-table sortable">
      <thead><tr>
        <th onclick="sortTable(this)">Sample Name</th>
        <th onclick="sortTable(this)">Level</th>
        <th onclick="sortTable(this)">Plate</th>
        <th onclick="sortTable(this)">Group</th>
        <th onclick="sortTable(this)">Avg Signal</th>
        <th onclick="sortTable(this)">{qc_hdr}</th>
        <th onclick="sortTable(this)">Expected Conc.</th>
        <th onclick="sortTable(this)">% Recovery</th>
      </tr></thead>
      <tbody>{''.join(qc_rows_html)}</tbody>
    </table>
    </div>"""

    # ── All Unknowns table rows (global sort + TP assignment matches create_output) ──
    # Build global unknowns dict keyed by (sample_name, group, plate) — same as create_output
    all_unk_groups = defaultdict(lambda: {'signals': [], 'concs': [], 'wells': [],
                                          'spot': None, 'group': '', 'plate': None,
                                          'lloq_sig': None, 'params': None, 'uloq_conc': None})
    for res in results:
        plate, spot, group = res['plate'], res['spot'], res.get('group', '')
        lloq_sig = res.get('lloq_sig')
        params = res['params']
        uloq_conc = lloq_conc_num = None
        if params is not None:
            try:
                stds = res.get('standards', [])
                uloq_conc = max((s['conc'] for s in stds if np.isfinite(s['conc'])), default=None)
                if lloq_sig is not None:
                    lc = inverse_4pl(lloq_sig, *params)
                    lloq_conc_num = lc if np.isfinite(lc) and lc > 0 else None
            except Exception:
                pass
        for u in res.get('unknowns', []):
            sname = u['sample_name']
            if _identify_qc_level(sname):
                continue
            key = (sname, group, plate)
            d = all_unk_groups[key]
            d['spot'] = spot; d['group'] = group; d['plate'] = plate
            d['lloq_sig'] = lloq_sig; d['params'] = params
            d['uloq_conc'] = uloq_conc; d['lloq_conc_num'] = lloq_conc_num
            if np.isfinite(u['signal']):
                d['signals'].append(u['signal'])
            if np.isfinite(u['interp_conc']):
                d['concs'].append(u['interp_conc'])
            d['wells'].append(u['well'])

    _sp_entries = []
    tp_index = defaultdict(int)
    unk_rows_html = []
    for (sname, group, plate), data in sorted(all_unk_groups.items()):
        spot = data['spot']
        lloq_sig = data['lloq_sig']
        uloq_conc = data.get('uloq_conc')
        lloq_conc_num = data.get('lloq_conc_num')

        avg_sig = np.mean(data['signals']) if data['signals'] else np.nan
        avg_conc = np.mean(data['concs']) if data['concs'] else np.nan
        cv = np.nan
        if len(data['concs']) > 1 and np.isfinite(avg_conc) and avg_conc != 0:
            cv = np.std(data['concs'], ddof=1) / avg_conc * 100

        flag = ''
        if np.isfinite(avg_sig) and lloq_sig is not None and avg_sig < lloq_sig:
            flag = '< LLOQ'
        elif np.isfinite(avg_conc):
            if uloq_conc and avg_conc > uloq_conc:
                flag = '> ULOQ'
            elif lloq_conc_num and avg_conc < lloq_conc_num:
                flag = '< LLOQ'
            else:
                flag = 'In Range'
        else:
            flag = 'Out of Range'

        # Dilution factor & corrected conc (QC > group > plate)
        factor = _resolve_dilution_factor(
            sname, group, plate,
            qc_dilution_factors, group_dilution_factors, plate_dilution_factors)
        corrected = avg_conc * factor if np.isfinite(avg_conc) else np.nan
        # Total protein & normalized (same in-order assignment as create_output)
        animal, tissue = _extract_animal_tissue(sname)
        tp_val = None
        if total_protein_map and animal:
            tp_key = (animal, tissue)
            tp_list = total_protein_map.get(tp_key, [])
            idx = tp_index[tp_key]
            if idx < len(tp_list):
                tp_val = tp_list[idx]
                tp_index[tp_key] += 1
        norm_val = (corrected / tp_val
                    if tp_val is not None and np.isfinite(corrected) and tp_val != 0
                    else None)

        if np.isfinite(corrected) and not _identify_qc_level(sname):
            _sp_entries.append({'analyte': group or 'Default', 'sample': sname,
                                'conc': float(corrected),
                                'norm': float(norm_val) if norm_val is not None else None,
                                'flag': flag, 'plate': plate})

        flag_class = ('status-good' if flag == 'In Range'
                      else 'status-warn' if flag in ('> ULOQ', '< LLOQ') else '')
        cv_class = 'cv-bad' if np.isfinite(cv) and cv > cv_threshold else ''
        avg_sig_str  = f"{avg_sig:,.1f}" if np.isfinite(avg_sig) else 'N/A'
        avg_conc_str = f"{avg_conc:.4g}" if np.isfinite(avg_conc) else 'N/A'
        cv_str       = f"{cv:.1f}" if np.isfinite(cv) else 'N/A'
        corr_str     = f"{corrected:.4g}" if np.isfinite(corrected) else ''
        factor_str   = str(factor) if factor and factor != 1.0 else ''
        tp_str       = f"{tp_val:.4g}" if tp_val is not None else ''
        norm_str     = f"{norm_val:.6g}" if norm_val is not None else ''
        animal_str   = animal or ''
        tissue_str   = tissue or ''
        unk_rows_html.append(
            f"<tr><td>{sname}</td><td>{animal_str}</td><td>{tissue_str}</td>"
            f"<td>{plate}</td><td>{spot}</td><td>{group}</td>"
            f"<td>{', '.join(data['wells'])}</td><td>{avg_sig_str}</td>"
            f"<td>{avg_conc_str}</td><td class='{cv_class}'>{cv_str}</td>"
            f"<td class='{flag_class}'>{flag}</td>"
            f"<td>{factor_str}</td><td>{corr_str}</td>"
            f"<td>{tp_str}</td><td>{norm_str}</td></tr>"
        )

    # ── Sample plot JSON ──────────────────────────────────────────────────────
    _sp_by_analyte = defaultdict(lambda: defaultdict(list))
    for e in _sp_entries:
        _sp_by_analyte[e['analyte']][e['sample']].append(
            {'conc': e['conc'], 'norm': e.get('norm'), 'flag': e['flag'], 'plate': e['plate']})

    _sp_has_norm = any(e.get('norm') is not None for e in _sp_entries)
    _sp_data = {'analytes': [], 'units': units or '', 'hasNorm': _sp_has_norm, 'samples': {}}
    for _sp_analyte in sorted(_sp_by_analyte.keys()):
        _sp_data['analytes'].append(_sp_analyte)
        _sp_data['samples'][_sp_analyte] = []
        for _sp_sname in sorted(_sp_by_analyte[_sp_analyte].keys()):
            _sp_elist = _sp_by_analyte[_sp_analyte][_sp_sname]
            _sp_concs = [_e['conc'] for _e in _sp_elist]
            _sp_norms = [_e['norm'] for _e in _sp_elist if _e.get('norm') is not None]
            _sp_flags = [_e['flag'] for _e in _sp_elist]
            _sp_mean = float(np.mean(_sp_concs))
            _sp_sd = float(np.std(_sp_concs, ddof=1)) if len(_sp_concs) > 1 else 0.0
            _sp_norm_mean = float(np.mean(_sp_norms)) if _sp_norms else None
            _sp_norm_sd = float(np.std(_sp_norms, ddof=1)) if len(_sp_norms) > 1 else (0.0 if _sp_norms else None)
            _sp_any_flagged = any(f != 'In Range' for f in _sp_flags)
            _sp_data['samples'][_sp_analyte].append({
                'name': _sp_sname, 'mean': _sp_mean, 'sd': _sp_sd, 'values': _sp_concs,
                'normMean': _sp_norm_mean, 'normSd': _sp_norm_sd, 'normValues': _sp_norms,
                'flags': _sp_flags, 'anyFlagged': _sp_any_flagged
            })
    _sp_json = _json.dumps(_sp_data)

    # ── QC plot JSON ──────────────────────────────────────────────────────────
    all_qc_groups = defaultdict(lambda: {'signals': [], 'concs': [], 'wells': [],
                                          'spot': None, 'group': '', 'plate': None})
    for res in results:
        plate, spot, group = res['plate'], res['spot'], res.get('group', '')
        for u in res.get('unknowns', []):
            sname = u['sample_name']
            if not _identify_qc_level(sname):
                continue
            key = (sname, group, plate)
            d = all_qc_groups[key]
            d['spot'] = spot; d['group'] = group; d['plate'] = plate
            if np.isfinite(u['signal']):
                d['signals'].append(u['signal'])
            if np.isfinite(u['interp_conc']):
                d['concs'].append(u['interp_conc'])
            d['wells'].append(u['well'])

    _qp_entries = []
    for (sname, group, plate), data in sorted(all_qc_groups.items()):
        avg_conc = np.mean(data['concs']) if data['concs'] else np.nan
        if not np.isfinite(avg_conc):
            continue
        level = _identify_qc_level(sname)
        qc_factor = 1.0
        if qc_dilution_factors:
            _hgrp = group if group and group != '_default' else ''
            _grp_qc = (qc_dilution_factors or {}).get(_hgrp, {})
            if level and level in _grp_qc:
                qc_factor = _grp_qc[level]
        corrected = avg_conc * qc_factor
        exp_conc = None
        if qc_expected_concentrations:
            _hgrp = group if group and group != '_default' else ''
            exp_conc = (qc_expected_concentrations.get(_hgrp)
                        if isinstance(qc_expected_concentrations, dict)
                        else qc_expected_concentrations)
        recovery = (corrected / exp_conc * 100) if (exp_conc and exp_conc != 0) else None
        _qp_entries.append({
            'analyte': group or 'Default', 'sample': sname, 'level': level or '',
            'plate': plate, 'conc': float(corrected),
            'expected': float(exp_conc) if exp_conc is not None else None,
            'recovery': float(recovery) if recovery is not None else None,
            'values': [float(c * qc_factor) for c in data['concs']],
        })

    _qp_by_analyte = defaultdict(lambda: defaultdict(list))
    for e in _qp_entries:
        _qp_by_analyte[e['analyte']][e['sample']].append(e)

    _qp_data = {'analytes': [], 'units': units or '', 'levels': list(QC_LEVELS), 'samples': {}, 'expected': {}}
    for _qp_analyte in sorted(_qp_by_analyte.keys()):
        _qp_data['analytes'].append(_qp_analyte)
        _qp_data['samples'][_qp_analyte] = []
        for _qp_sname in sorted(_qp_by_analyte[_qp_analyte].keys()):
            _qp_elist = _qp_by_analyte[_qp_analyte][_qp_sname]
            _qp_concs = [e['conc'] for e in _qp_elist]
            _qp_vals = [v for e in _qp_elist for v in e.get('values', [])]
            _qp_mean = float(np.mean(_qp_concs))
            _qp_sd = float(np.std(_qp_concs, ddof=1)) if len(_qp_concs) > 1 else 0.0
            _qp_level = _qp_elist[0]['level']
            _qp_exp = _qp_elist[0].get('expected')
            if _qp_exp is not None:
                _qp_data['expected'][_qp_analyte] = _qp_exp
            _qp_data['samples'][_qp_analyte].append({
                'name': _qp_sname, 'level': _qp_level,
                'mean': _qp_mean, 'sd': _qp_sd, 'values': _qp_vals,
            })
    _qp_json = _json.dumps(_qp_data)

    # ── Assemble curve cards HTML ─────────────────────────────────────────────
    curves_section_html = '\n'.join(
        f'<div class="curve-card"><h3>{label}</h3>{div_html}</div>'
        for label, div_html in curve_divs
    )

    # ── Plotly JS bundle (self-contained) ─────────────────────────────────────
    plotly_js = poff.get_plotlyjs()

    msd_basename = os.path.basename(msd_path)
    excel_basename = os.path.basename(excel_path) if excel_path else None
    excel_abs = ('file://' + os.path.abspath(excel_path).replace('\\', '/')) if excel_path else None
    excel_btn_html = (
        f'<a class="excel-btn" href="{excel_abs}">⬇ Open Excel</a>'
        if excel_abs else ''
    )
    has_tp = bool(total_protein_map)
    tp_headers = (
        "<th onclick=\"sortTable(this)\">Total Protein</th>"
        "<th onclick=\"sortTable(this)\">Normalized Conc.</th>"
    ) if has_tp else ""
    unk_hdr_row = (
        "<tr>"
        "<th onclick=\"sortTable(this)\">Sample Name</th>"
        "<th onclick=\"sortTable(this)\">Animal</th>"
        "<th onclick=\"sortTable(this)\">Tissue</th>"
        "<th onclick=\"sortTable(this)\">Plate</th>"
        "<th onclick=\"sortTable(this)\">Spot</th>"
        "<th onclick=\"sortTable(this)\">Group</th>"
        "<th onclick=\"sortTable(this)\">Wells</th>"
        "<th onclick=\"sortTable(this)\">Avg Signal</th>"
        f"<th onclick=\"sortTable(this)\">Avg Interp. Conc.{unit_suffix}</th>"
        "<th onclick=\"sortTable(this)\">%CV</th>"
        "<th onclick=\"sortTable(this)\">Flag</th>"
        "<th onclick=\"sortTable(this)\">Dilution Factor</th>"
        f"<th onclick=\"sortTable(this)\">Corrected Avg Conc.{unit_suffix}</th>"
        + tp_headers +
        "</tr>"
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>MSD 4PL Analysis Report</title>
<script>{plotly_js}</script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 13px; background: #f0f2f5; color: #222; }}
  .header {{ background: #3a506b; color: white; padding: 18px 28px; }}
  .header h1 {{ font-size: 22px; font-weight: bold; letter-spacing: 0.5px; }}
  .header p {{ font-size: 12px; opacity: 0.8; margin-top: 4px; }}
  .header .accent {{ height: 3px; background: #7ba7bc; margin-top: 10px; border-radius: 2px; }}
  .tabs {{ display: flex; background: #2e3f52; padding: 0 20px; }}
  .tab-btn {{ padding: 12px 22px; cursor: pointer; color: #c5d5e8; border: none;
              background: none; font-size: 13px; font-weight: 500;
              border-bottom: 3px solid transparent; }}
  .tab-btn:hover {{ color: white; }}
  .tab-btn.active {{ color: white; border-bottom-color: #7ba7bc; }}
  .content {{ padding: 24px 28px; max-width: 1400px; margin: 0 auto; }}
  .tab-pane {{ display: none; }}
  .tab-pane.active {{ display: block; }}
  .table-wrap {{ overflow-x: auto; margin-bottom: 24px; }}
  .data-table {{ border-collapse: collapse; width: 100%; background: white;
                  box-shadow: 0 1px 4px rgba(0,0,0,0.08); border-radius: 4px; }}
  .data-table th {{ background: #2F5496; color: white; padding: 9px 12px;
                    text-align: left; cursor: pointer; white-space: nowrap; user-select: none; }}
  .data-table th:hover {{ background: #3a65b5; }}
  .data-table th.sort-asc::after {{ content: ' ▲'; font-size: 10px; }}
  .data-table th.sort-desc::after {{ content: ' ▼'; font-size: 10px; }}
  .data-table td {{ padding: 7px 12px; border-bottom: 1px solid #e8edf3; vertical-align: middle; }}
  .data-table tr:hover td {{ background: #f5f8ff; }}
  .data-table tr:last-child td {{ border-bottom: none; }}
  .status-good {{ color: #006100; font-weight: 500; }}
  .status-warn {{ color: #9C5700; font-weight: 500; }}
  .status-fail {{ color: #9C0006; font-weight: 500; }}
  .cv-bad {{ background: #F8CBAD; }}
  .curves-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(520px, 1fr)); gap: 20px; }}
  .export-bar {{ display: flex; justify-content: flex-end; margin-bottom: 10px; }}
  .export-btn {{ padding: 8px 18px; background: #3a506b; color: white; border: none;
                 border-radius: 4px; font-size: 13px; cursor: pointer; font-weight: 500; }}
  .export-btn:hover {{ background: #2e3f52; }}
  .excel-btn {{ padding: 8px 18px; background: #1e6b3c; color: white; border: none;
                border-radius: 4px; font-size: 13px; cursor: pointer; font-weight: 500;
                text-decoration: none; display: inline-block; }}
  .excel-btn:hover {{ background: #155230; }}
  @media print {{
    .tabs, .export-bar {{ display: none !important; }}
    .tab-pane {{ display: block !important; page-break-inside: avoid; }}
    .tab-pane + .tab-pane {{ page-break-before: always; }}
    body {{ background: white; font-size: 11px; }}
    .content {{ max-width: 100%; padding: 10px; }}
    .data-table th {{ background: #2F5496 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .curve-card {{ box-shadow: none; border: 1px solid #ccc; }}
    .section {{ box-shadow: none; border: 1px solid #ccc; }}
    .header {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
  }}
  .curve-card {{ background: white; border-radius: 6px; box-shadow: 0 1px 4px rgba(0,0,0,0.08);
                  padding: 16px; }}
  .curve-card h3 {{ font-size: 13px; color: #3a506b; margin-bottom: 10px; font-weight: 600; }}
  h2 {{ font-size: 16px; color: #3a506b; margin: 20px 0 12px; font-weight: 700; }}
  .section {{ background: white; border-radius: 6px; box-shadow: 0 1px 4px rgba(0,0,0,0.08);
               padding: 20px; margin-bottom: 24px; }}
  .curve-toggle-btn {{ padding: 5px 14px; border: none; border-radius: 4px; cursor: pointer;
                       font-size: 12px; font-weight: 600; background: #27AE60; color: white;
                       margin-bottom: 6px; display: inline-block; }}
  .curve-toggle-btn:hover {{ background: #1e8449; }}
  .sp-panel {{ background:white;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,0.08);padding:14px; }}
  .sp-drop-zone {{ min-height:80px;border:2px dashed #ccc;border-radius:4px;padding:6px;display:flex;flex-wrap:wrap;gap:4px;align-content:flex-start;transition:background 0.15s; }}
  .sp-drop-zone.drag-over {{ background:#e8f4fd;border-color:#2F5496; }}
  .sp-chip {{ display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:12px;font-size:11px;cursor:grab;user-select:none;border:1px solid rgba(0,0,0,0.15);background:white; }}
  .sp-chip.flagged {{ border-color:#c0392b;color:#c0392b; }}
  .sp-chip input[type=checkbox] {{ cursor:pointer;margin:0; }}
  .sp-group-block {{ margin-bottom:10px;border-radius:6px;overflow:hidden;border:1px solid #ddd; }}
  .sp-group-header {{ display:flex;align-items:center;gap:6px;padding:6px 10px;font-size:12px;font-weight:600;color:white; }}
  .sp-group-drop {{ min-height:36px;padding:6px;display:flex;flex-wrap:wrap;gap:4px;align-content:flex-start; }}
  .sp-btn {{ padding:5px 12px;border:1px solid #ccc;border-radius:4px;cursor:pointer;font-size:12px;background:white;color:#333; }}
  .sp-btn:hover {{ background:#f0f0f0; }}
  .sp-btn-primary {{ background:#2F5496;color:white;border-color:#2F5496; }}
  .sp-btn-primary:hover {{ background:#3a65b5; }}
  .sp-btn-icon {{ padding:3px 7px;font-size:11px;border-radius:3px; }}
  .sp-sort-btn {{ background:white; }}
  .active-sort {{ background:#2F5496 !important;color:white !important;border-color:#2F5496 !important; }}
  .sp-analyte-btn {{ padding:7px 18px;border:none;border-radius:4px;cursor:pointer;font-size:13px;font-weight:500;background:#dde3ec;color:#333; }}
  .sp-analyte-btn.active {{ background:#2F5496;color:white; }}
</style>
</head>
<body>

<div class="header">
  <h1>MSD 4PL Analysis Report</h1>
  <p>{msd_basename}</p>
  <div class="accent"></div>
</div>

<div class="tabs">
  <button class="tab-btn active" onclick="showTab('summary', this)">Summary</button>
  <button class="tab-btn" onclick="showTab('curves', this)">Standard Curves</button>
  <button class="tab-btn" onclick="showTab('unknowns', this)">All Unknowns</button>
  <button class="tab-btn" onclick="showTab('sampleplots', this)">Sample Plots</button>
  <button class="tab-btn" onclick="showTab('qcplots', this)">QC Plots</button>
  <div style="margin-left:auto;display:flex;align-items:center;gap:8px;padding-right:12px;">
    {excel_btn_html}
    <button class="export-btn" onclick="window.print()">⬇ Export PDF</button>
  </div>
</div>

<div class="content">

  <div id="tab-summary" class="tab-pane active">
    <h2>Curve Fit Summary</h2>
    <p style="font-size:12px;color:#555;margin:-8px 0 12px;"><strong>LLOQ Method:</strong> {lloq_method_label}</p>
    <div class="table-wrap">
    <table id="summaryTable" class="data-table">
      <thead><tr>
        <th onclick="sortTable(this)">Plate</th>
        <th onclick="sortTable(this)">Spot</th>
        <th onclick="sortTable(this)">Group</th>
        <th onclick="sortTable(this)">Min (a)</th>
        <th onclick="sortTable(this)">Hill Slope (b)</th>
        <th onclick="sortTable(this)">EC50 (c)</th>
        <th onclick="sortTable(this)">Max (d)</th>
        <th onclick="sortTable(this)">LLOQ Signal</th>
        <th onclick="sortTable(this)">LLOQ Conc</th>
        <th onclick="sortTable(this)">R²</th>
        <th onclick="sortTable(this)">Flags</th>
        <th onclick="sortTable(this)">Status</th>
      </tr></thead>
      <tbody>{''.join(summary_rows_html)}</tbody>
    </table>
    </div>
    {qc_table_html}
    <h2>Standard Curve Overlay</h2>
    <div class="section">
      {_overlay_btns}
      {overlay_div}
    </div>
  </div>

  <div id="tab-curves" class="tab-pane">
    <h2>Standard Curves</h2>
    <div class="curves-grid">
      {curves_section_html}
    </div>
  </div>

  <div id="tab-unknowns" class="tab-pane">
    <h2>All Unknowns</h2>
    <div class="table-wrap">
    <table id="unkTable" class="data-table">
      <thead>{unk_hdr_row}</thead>
      <tbody>{''.join(unk_rows_html)}</tbody>
    </table>
    </div>
  </div>

  <div id="tab-sampleplots" class="tab-pane">
    <h2>Sample Plots</h2>
    <div id="sp-analyte-bar" style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px;"></div>
    <div style="display:grid;grid-template-columns:280px 1fr;gap:16px;margin-bottom:16px;">
      <div style="display:flex;flex-direction:column;gap:10px;">
        <div class="sp-panel">
          <div style="font-weight:600;font-size:13px;color:#3a506b;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;">
            Unassigned Samples
            <span style="display:flex;gap:8px;align-items:center;">
              <button class="sp-btn sp-btn-icon" id="sp-select-all-btn" onclick="spSelectAllUnassigned(this)" style="font-size:11px;">Select All</button>
              <label style="font-weight:400;font-size:12px;cursor:pointer;">
                <input type="checkbox" id="sp-show-unassigned" checked onchange="spRenderChart()"> Show
              </label>
            </span>
          </div>
          <div id="sp-unassigned-pool" class="sp-drop-zone" ondragover="spDragOver(event)" ondrop="spDrop(event,'__unassigned__')"></div>
          <div style="margin-top:8px;display:flex;gap:6px;align-items:center;flex-wrap:wrap;">
            <button class="sp-btn sp-btn-primary" onclick="spAssignChecked()">Assign to Group &#x2192;</button>
            <button class="sp-btn" onclick="spCreateGroup()">&#xFF0B; New Group</button>
          </div>
        </div>
        <div class="sp-panel" style="flex:1;">
          <div style="font-weight:600;font-size:13px;color:#3a506b;margin-bottom:8px;">Groups</div>
          <div id="sp-groups-container"></div>
        </div>
      </div>
      <div>
        <div id="sp-value-toggle" style="display:flex;gap:8px;align-items:center;margin-bottom:10px;flex-wrap:wrap;">
          <span style="font-size:12px;font-weight:600;color:#555;">Values:</span>
          <button class="sp-btn sp-sort-btn active-sort" id="sp-val-corrected" onclick="spSetValueMode('corrected',this)">Corrected Conc.</button>
          <button class="sp-btn sp-sort-btn" id="sp-val-norm" onclick="spSetValueMode('normalized',this)" title="Requires total protein data">Normalized Protein</button>
        </div>
        <div style="display:flex;gap:8px;align-items:center;margin-bottom:10px;flex-wrap:wrap;">
          <span style="font-size:12px;font-weight:600;color:#555;">Sort:</span>
          <button class="sp-btn sp-sort-btn active-sort" id="sp-sort-group" onclick="spSetSort('group',this)">By Group</button>
          <button class="sp-btn sp-sort-btn" id="sp-sort-asc" onclick="spSetSort('asc',this)">Value &#x2191;</button>
          <button class="sp-btn sp-sort-btn" id="sp-sort-desc" onclick="spSetSort('desc',this)">Value &#x2193;</button>
        </div>
        <div id="sp-chart" style="width:100%;"></div>
      </div>
    </div>
  </div>

  <div id="tab-qcplots" class="tab-pane">
    <h2>QC Plots</h2>
    <div id="qp-analyte-bar" style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px;"></div>
    <div style="display:grid;grid-template-columns:280px 1fr;gap:16px;margin-bottom:16px;">
      <div style="display:flex;flex-direction:column;gap:10px;">
        <div class="sp-panel">
          <div style="font-weight:600;font-size:13px;color:#3a506b;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;">
            Unassigned Samples
            <span style="display:flex;gap:8px;align-items:center;">
              <button class="sp-btn sp-btn-icon" id="qp-select-all-btn" onclick="qpSelectAllUnassigned(this)" style="font-size:11px;">Select All</button>
              <label style="font-weight:400;font-size:12px;cursor:pointer;">
                <input type="checkbox" id="qp-show-unassigned" checked onchange="qpRenderChart()"> Show
              </label>
            </span>
          </div>
          <div id="qp-unassigned-pool" class="sp-drop-zone" ondragover="qpDragOver(event)" ondrop="qpDrop(event,'__unassigned__')"></div>
          <div style="margin-top:8px;display:flex;gap:6px;align-items:center;flex-wrap:wrap;">
            <button class="sp-btn sp-btn-primary" onclick="qpAssignChecked()">Assign to Group &#x2192;</button>
            <button class="sp-btn" onclick="qpCreateGroup()">&#xFF0B; New Group</button>
          </div>
        </div>
        <div class="sp-panel" style="flex:1;">
          <div style="font-weight:600;font-size:13px;color:#3a506b;margin-bottom:8px;">Groups</div>
          <div id="qp-groups-container"></div>
        </div>
      </div>
      <div>
        <div style="display:flex;gap:8px;align-items:center;margin-bottom:10px;flex-wrap:wrap;">
          <span style="font-size:12px;font-weight:600;color:#555;">Sort:</span>
          <button class="sp-btn sp-sort-btn active-sort" id="qp-sort-group" onclick="qpSetSort('group',this)">By Group</button>
          <button class="sp-btn sp-sort-btn" id="qp-sort-asc" onclick="qpSetSort('asc',this)">Value &#x2191;</button>
          <button class="sp-btn sp-sort-btn" id="qp-sort-desc" onclick="qpSetSort('desc',this)">Value &#x2193;</button>
        </div>
        <div id="qp-chart" style="width:100%;"></div>
      </div>
    </div>
  </div>

</div>

<script>
function showTab(name, btn) {{
  document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  const pane = document.getElementById('tab-' + name);
  pane.classList.add('active');
  btn.classList.add('active');
  // Resize all Plotly charts now that their containers are visible
  pane.querySelectorAll('.js-plotly-plot').forEach(el => Plotly.Plots.resize(el));
  if (name === 'sampleplots') spInit();
  if (name === 'qcplots') qpInit();
}}

function sortTable(th) {{
  const table = th.closest('table');
  const tbody = table.querySelector('tbody');
  const col = Array.from(th.parentNode.children).indexOf(th);
  const asc = th.classList.contains('sort-asc');
  table.querySelectorAll('th').forEach(h => h.classList.remove('sort-asc', 'sort-desc'));
  th.classList.add(asc ? 'sort-desc' : 'sort-asc');
  const dir = asc ? -1 : 1;
  const rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a, b) => {{
    const av = a.cells[col]?.textContent.trim() ?? '';
    const bv = b.cells[col]?.textContent.trim() ?? '';
    const an = parseFloat(av.replace(/[,%]/g, ''));
    const bn = parseFloat(bv.replace(/[,%]/g, ''));
    if (!isNaN(an) && !isNaN(bn)) return (an - bn) * dir;
    return av.localeCompare(bv) * dir;
  }});
  rows.forEach(r => tbody.appendChild(r));
}}

function filterTable(query, tableId) {{
  const q = query.toLowerCase().trim();
  document.getElementById(tableId).querySelectorAll('tbody tr').forEach(row => {{
    const match = !q || row.textContent.toLowerCase().includes(q);
    row.style.display = match ? 'table-row' : 'none';
  }});
}}

function msdToggleGrp(btn, traceIndices, shapeIndices) {{
  var gd = document.getElementById('overlay_chart');
  var active = btn.getAttribute('data-active') === '1';
  if (traceIndices && traceIndices.length) {{
    Plotly.restyle(gd, {{visible: active ? false : true}}, traceIndices);
  }}
  if (shapeIndices && shapeIndices.length) {{
    var shapeUpd = {{}};
    shapeIndices.forEach(function(i) {{ shapeUpd['shapes[' + i + '].visible'] = !active; }});
    Plotly.relayout(gd, shapeUpd);
  }}
  btn.setAttribute('data-active', active ? '0' : '1');
  btn.style.opacity = active ? '0.4' : '1.0';
}}
function msdOverlayAll(show) {{
  var gd = document.getElementById('overlay_chart');
  var allTraceIdx = {_json.dumps(_all_grp_indices)};
  var allShapeIdx = {_json.dumps(_all_shape_indices)};
  if (allTraceIdx.length) {{
    Plotly.restyle(gd, {{visible: show ? true : false}}, allTraceIdx);
  }}
  if (allShapeIdx.length) {{
    var shapeUpd = {{}};
    allShapeIdx.forEach(function(i) {{ shapeUpd['shapes[' + i + '].visible'] = show; }});
    Plotly.relayout(gd, shapeUpd);
  }}
  document.querySelectorAll('[data-active]').forEach(function(b) {{
    b.setAttribute('data-active', show ? '1' : '0');
    b.style.opacity = show ? '1.0' : '0.4';
  }});
}}
function msdToggleCurveSamples(btn, divId, traceIdx) {{
  var gd = document.getElementById(divId);
  var active = btn.getAttribute('data-active') === '1';
  Plotly.restyle(gd, {{visible: active ? false : true}}, [traceIdx]);
  btn.setAttribute('data-active', active ? '0' : '1');
  btn.style.opacity = active ? '0.4' : '1.0';
}}

// ── Sample Plots Tab ─────────────────────────────────────────────────────────
var SP_DATA = {_sp_json};
var spInitialized = false;
var spCurrentAnalyte = null;
var spGroups = [];          // [{{id, name, color, visible, samples:[]}}]
var spUnassigned = [];      // [sampleName, ...]
var spSortMode = 'group';
var spGroupIdCounter = 0;
var spDragPayload = null;   // {{name, fromGroup}}
var spLastCheckedIdx = -1;  // for shift+click range selection in unassigned pool
var spValueMode = 'corrected';  // 'corrected' | 'normalized'
var SP_PALETTE = ['#1f77b4','#ff7f0e','#2ca02c','#9467bd','#8c564b','#e377c2','#17becf','#bcbd22'];

var QP_DATA = {_qp_json};
var qpInitialized = false;
var qpCurrentAnalyte = null;
var qpGroups = [];
var qpUnassigned = [];
var qpSortMode = 'group';
var qpGroupIdCounter = 0;
var qpDragPayload = null;
var qpLastCheckedIdx = -1;
var QP_PALETTE = ['#e41a1c','#ff7f00','#4daf4a','#377eb8','#984ea3','#a65628','#f781bf','#999999'];

function spNextColor() {{
  return SP_PALETTE[spGroups.length % SP_PALETTE.length];
}}

function spInit() {{
  if (!spInitialized) {{
    spInitialized = true;
    spGroups = [];
    spGroupIdCounter = 0;
    spSortMode = 'group';
    spValueMode = 'corrected';
    var analytes = SP_DATA.analytes || [];
    spCurrentAnalyte = analytes.length > 0 ? analytes[0] : null;
    // Populate unassigned with all samples for current analyte
    spUnassigned = spCurrentAnalyte && SP_DATA.samples[spCurrentAnalyte]
      ? SP_DATA.samples[spCurrentAnalyte].map(function(s) {{ return s.name; }})
      : [];
    spBuildAnalyteBar();
    // Show value-mode toggle only when normalized data is available
    var normBtn = document.getElementById('sp-val-norm');
    if (normBtn && !SP_DATA.hasNorm) {{
      normBtn.disabled = true;
      normBtn.style.opacity = '0.4';
      normBtn.style.cursor = 'not-allowed';
      normBtn.title = 'No total protein data loaded';
    }}
  }}
  spRenderGroupPanel();
  spRenderChart();
}}

function spBuildAnalyteBar() {{
  var bar = document.getElementById('sp-analyte-bar');
  if (!bar) return;
  bar.innerHTML = '';
  (SP_DATA.analytes || []).forEach(function(a) {{
    var btn = document.createElement('button');
    btn.className = 'sp-analyte-btn' + (a === spCurrentAnalyte ? ' active' : '');
    btn.textContent = a;
    btn.onclick = function() {{ spSelectAnalyte(a); }};
    bar.appendChild(btn);
  }});
}}

function spSelectAnalyte(name) {{
  spCurrentAnalyte = name;
  // Update unassigned: samples in current analyte not already in a group
  var allSamples = SP_DATA.samples[name] ? SP_DATA.samples[name].map(function(s) {{ return s.name; }}) : [];
  var inGroups = {{}};
  spGroups.forEach(function(g) {{ g.samples.forEach(function(s) {{ inGroups[s] = true; }}); }});
  spUnassigned = allSamples.filter(function(s) {{ return !inGroups[s]; }});
  document.querySelectorAll('.sp-analyte-btn').forEach(function(b) {{
    b.classList.toggle('active', b.textContent === name);
  }});
  spRenderGroupPanel();
  spRenderChart();
}}

function spGetSampleData(sname) {{
  if (!spCurrentAnalyte || !SP_DATA.samples[spCurrentAnalyte]) return null;
  var arr = SP_DATA.samples[spCurrentAnalyte];
  for (var i = 0; i < arr.length; i++) {{
    if (arr[i].name === sname) return arr[i];
  }}
  return null;
}}

function spRenderGroupPanel() {{
  spLastCheckedIdx = -1;  // reset range anchor whenever the pool is rebuilt
  // Unassigned pool
  var pool = document.getElementById('sp-unassigned-pool');
  if (pool) {{
    pool.innerHTML = '';
    spUnassigned.forEach(function(sname) {{
      pool.appendChild(spMakeChip(sname, '__unassigned__'));
    }});
  }}
  // Groups container
  var gc = document.getElementById('sp-groups-container');
  if (!gc) return;
  gc.innerHTML = '';
  spGroups.forEach(function(g) {{
    var block = document.createElement('div');
    block.className = 'sp-group-block';
    // Header
    var hdr = document.createElement('div');
    hdr.className = 'sp-group-header';
    hdr.style.background = g.color;
    var isFirst = (spGroups.indexOf(g) === 0);
    var isLast  = (spGroups.indexOf(g) === spGroups.length - 1);
    var btnStyle = 'background:rgba(255,255,255,0.25);color:white;border-color:rgba(255,255,255,0.4);';
    var btnDisabled = 'background:rgba(255,255,255,0.08);color:rgba(255,255,255,0.3);border-color:rgba(255,255,255,0.15);cursor:default;';
    hdr.innerHTML =
      '<span style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="' + g.name + '">' + g.name + '</span>' +
      '<button class="sp-btn sp-btn-icon" style="' + (isFirst ? btnDisabled : btnStyle) + '" ' +
        (isFirst ? 'disabled ' : 'onclick="spMoveGroup(' + g.id + ',-1)" ') + 'title="Move up">&#x25B4;</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + (isLast  ? btnDisabled : btnStyle) + '" ' +
        (isLast  ? 'disabled ' : 'onclick="spMoveGroup(' + g.id + ',1)" ')  + 'title="Move down">&#x25BE;</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + btnStyle + '" ' +
        'onclick="spRenameGroup(' + g.id + ')" title="Rename group">&#x270F;</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + btnStyle + '" ' +
        'onclick="spToggleGroup(' + g.id + ')" title="Show/hide in chart">' + (g.visible ? '&#128065;' : '&#128564;') + '</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + btnStyle + '" ' +
        'onclick="spDeleteGroup(' + g.id + ')" title="Delete group">&#x2715;</button>';
    block.appendChild(hdr);
    // Drop zone
    var dz = document.createElement('div');
    dz.className = 'sp-group-drop sp-drop-zone';
    dz.setAttribute('ondragover', 'spDragOver(event)');
    dz.setAttribute('ondrop', "spDrop(event,'" + g.id + "')");
    g.samples.forEach(function(sname) {{
      dz.appendChild(spMakeChip(sname, g.id));
    }});
    block.appendChild(dz);
    gc.appendChild(block);
  }});
}}

function spMakeChip(sname, groupId) {{
  var d = spGetSampleData(sname);
  var flagged = d && d.anyFlagged;
  var span = document.createElement('span');
  span.className = 'sp-chip' + (flagged ? ' flagged' : '');
  span.draggable = true;
  span.title = sname;
  var label = (flagged ? '⚠ ' : '') + sname;
  if (groupId === '__unassigned__') {{
    // Add checkbox with shift+click range selection
    var cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.setAttribute('data-sample', sname);
    cb.onclick = function(e) {{
      e.stopPropagation();
      var idx = spUnassigned.indexOf(sname);
      var allCbs = Array.from(document.querySelectorAll('#sp-unassigned-pool input[type=checkbox]'));
      if (e.shiftKey && spLastCheckedIdx >= 0 && idx !== spLastCheckedIdx) {{
        var lo = Math.min(spLastCheckedIdx, idx);
        var hi = Math.max(spLastCheckedIdx, idx);
        var newState = allCbs[idx].checked;
        for (var i = lo; i <= hi; i++) {{
          if (allCbs[i]) allCbs[i].checked = newState;
        }}
      }}
      spLastCheckedIdx = idx;
    }};
    span.appendChild(cb);
  }}
  var txt = document.createElement('span');
  txt.textContent = label;
  txt.style.maxWidth = '120px';
  txt.style.overflow = 'hidden';
  txt.style.textOverflow = 'ellipsis';
  txt.style.whiteSpace = 'nowrap';
  span.appendChild(txt);
  span.addEventListener('dragstart', function(e) {{ spDragStart(e, sname, groupId); }});
  span.addEventListener('dragend', function(e) {{ spDragEnd(e); }});
  return span;
}}

function spDragStart(e, name, fromGroup) {{
  spDragPayload = {{name: name, fromGroup: fromGroup}};
  e.dataTransfer.effectAllowed = 'move';
  e.currentTarget.style.opacity = '0.4';
}}

function spDragEnd(e) {{
  e.currentTarget.style.opacity = '1';
  document.querySelectorAll('.sp-drop-zone').forEach(function(z) {{
    z.classList.remove('drag-over');
  }});
}}

function spDragOver(e) {{
  e.preventDefault();
  e.dataTransfer.dropEffect = 'move';
  var zone = e.currentTarget.closest('.sp-drop-zone');
  if (zone) zone.classList.add('drag-over');
}}

function spDrop(e, toGroupId) {{
  e.preventDefault();
  document.querySelectorAll('.sp-drop-zone').forEach(function(z) {{ z.classList.remove('drag-over'); }});
  if (!spDragPayload) return;
  var name = spDragPayload.name;
  var fromGroup = spDragPayload.fromGroup;
  spDragPayload = null;
  if (fromGroup === toGroupId) return;
  // Remove from source
  if (fromGroup === '__unassigned__') {{
    spUnassigned = spUnassigned.filter(function(s) {{ return s !== name; }});
  }} else {{
    var sg = spGroups.find(function(g) {{ return g.id == fromGroup; }});
    if (sg) sg.samples = sg.samples.filter(function(s) {{ return s !== name; }});
  }}
  // Add to destination
  if (toGroupId === '__unassigned__') {{
    if (spUnassigned.indexOf(name) === -1) spUnassigned.push(name);
  }} else {{
    var tg = spGroups.find(function(g) {{ return g.id == toGroupId; }});
    if (tg && tg.samples.indexOf(name) === -1) tg.samples.push(name);
  }}
  spRenderGroupPanel();
  spRenderChart();
}}

function spCreateGroup() {{
  var name = prompt('Group name:');
  if (!name || !name.trim()) return;
  name = name.trim();
  spGroups.push({{
    id: ++spGroupIdCounter,
    name: name,
    color: spNextColor(),
    visible: true,
    samples: []
  }});
  spRenderGroupPanel();
  spRenderChart();
}}

function spAssignChecked() {{
  var checked = [];
  document.querySelectorAll('#sp-unassigned-pool input[type=checkbox]:checked').forEach(function(cb) {{
    checked.push(cb.getAttribute('data-sample'));
  }});
  if (!checked.length) {{ alert('Check at least one sample first.'); return; }}
  var name = prompt('Assign to group name:');
  if (!name || !name.trim()) return;
  name = name.trim();
  var g = spGroups.find(function(x) {{ return x.name === name; }});
  if (!g) {{
    g = {{ id: ++spGroupIdCounter, name: name, color: spNextColor(), visible: true, samples: [] }};
    spGroups.push(g);
  }}
  checked.forEach(function(s) {{
    spUnassigned = spUnassigned.filter(function(u) {{ return u !== s; }});
    if (g.samples.indexOf(s) === -1) g.samples.push(s);
  }});
  spRenderGroupPanel();
  spRenderChart();
}}

function spMoveGroup(id, dir) {{
  var idx = spGroups.findIndex(function(x) {{ return x.id == id; }});
  var newIdx = idx + dir;
  if (newIdx < 0 || newIdx >= spGroups.length) return;
  var tmp = spGroups[idx];
  spGroups[idx] = spGroups[newIdx];
  spGroups[newIdx] = tmp;
  spRenderGroupPanel();
  spRenderChart();
}}
function spRenameGroup(id) {{
  var g = spGroups.find(function(x) {{ return x.id == id; }});
  if (!g) return;
  var newName = prompt('Rename group:', g.name);
  if (newName === null || newName.trim() === '') return;
  g.name = newName.trim();
  spRenderGroupPanel();
  spRenderChart();
}}
function spDeleteGroup(id) {{
  var g = spGroups.find(function(x) {{ return x.id == id; }});
  if (!g) return;
  // Return samples to unassigned
  g.samples.forEach(function(s) {{
    if (spUnassigned.indexOf(s) === -1) spUnassigned.push(s);
  }});
  spGroups = spGroups.filter(function(x) {{ return x.id != id; }});
  spRenderGroupPanel();
  spRenderChart();
}}

function spToggleGroup(id) {{
  var g = spGroups.find(function(x) {{ return x.id == id; }});
  if (!g) return;
  g.visible = !g.visible;
  spRenderGroupPanel();
  spRenderChart();
}}

function spSetSort(mode, btn) {{
  spSortMode = mode;
  document.querySelectorAll('.sp-sort-btn').forEach(function(b) {{ b.classList.remove('active-sort'); }});
  btn.classList.add('active-sort');
  spRenderChart();
}}

function spSelectAllUnassigned(btn) {{
  var cbs = document.querySelectorAll('#sp-unassigned-pool input[type=checkbox]');
  var allChecked = Array.from(cbs).every(function(cb) {{ return cb.checked; }});
  cbs.forEach(function(cb) {{ cb.checked = !allChecked; }});
  btn.textContent = allChecked ? 'Select All' : 'Deselect All';
  spLastCheckedIdx = -1;  // reset range anchor after bulk action
}}
function spSetValueMode(mode, btn) {{
  spValueMode = mode;
  ['sp-val-corrected','sp-val-norm'].forEach(function(id) {{
    var el = document.getElementById(id);
    if (el) el.classList.remove('active-sort');
  }});
  if (btn) btn.classList.add('active-sort');
  spRenderChart();
}}

// Helper: pick mean/sd/values from a data entry based on current value mode
function spGetVals(d) {{
  if (spValueMode === 'normalized' && d.normMean !== null && d.normMean !== undefined) {{
    return {{ mean: d.normMean, sd: d.normSd || 0, values: d.normValues || [] }};
  }}
  return {{ mean: d.mean, sd: d.sd || 0, values: d.values || [] }};
}}

function spRenderChart() {{
  if (!spCurrentAnalyte || !SP_DATA.samples[spCurrentAnalyte]) {{
    Plotly.purge('sp-chart');
    return;
  }}
  var allData = SP_DATA.samples[spCurrentAnalyte];
  var showUnassigned = document.getElementById('sp-show-unassigned') ? document.getElementById('sp-show-unassigned').checked : true;
  var units = SP_DATA.units || '';
  var yTitle = spValueMode === 'normalized'
    ? spCurrentAnalyte + ' Normalized Concentration'
    : spCurrentAnalyte + ' Concentration' + (units ? ' (' + units + ')' : '');

  // Build ordered list of {{sname, color, groupName}} segments
  var segments = [];  // [{{groupName, color, items:[sampleName]}}]

  spGroups.forEach(function(g) {{
    if (!g.visible) return;
    var items = g.samples.filter(function(s) {{
      return allData.some(function(d) {{ return d.name === s; }});
    }});
    if (spSortMode === 'asc') {{
      items.sort(function(a, b) {{
        var da = allData.find(function(d) {{ return d.name === a; }});
        var db = allData.find(function(d) {{ return d.name === b; }});
        return (da ? spGetVals(da).mean : 0) - (db ? spGetVals(db).mean : 0);
      }});
    }} else if (spSortMode === 'desc') {{
      items.sort(function(a, b) {{
        var da = allData.find(function(d) {{ return d.name === a; }});
        var db = allData.find(function(d) {{ return d.name === b; }});
        return (db ? spGetVals(db).mean : 0) - (da ? spGetVals(da).mean : 0);
      }});
    }}
    if (items.length) segments.push({{groupName: g.name, color: g.color, items: items, collapsed: true}});
  }});

  var unassignedItems = showUnassigned
    ? spUnassigned.filter(function(s) {{ return allData.some(function(d) {{ return d.name === s; }}); }})
    : [];
  if (spSortMode === 'asc') {{
    unassignedItems.sort(function(a, b) {{
      var da = allData.find(function(d) {{ return d.name === a; }});
      var db = allData.find(function(d) {{ return d.name === b; }});
      return (da ? spGetVals(da).mean : 0) - (db ? spGetVals(db).mean : 0);
    }});
  }} else if (spSortMode === 'desc') {{
    unassignedItems.sort(function(a, b) {{
      var da = allData.find(function(d) {{ return d.name === a; }});
      var db = allData.find(function(d) {{ return d.name === b; }});
      return (db ? spGetVals(db).mean : 0) - (da ? spGetVals(da).mean : 0);
    }});
  }}
  if (unassignedItems.length) segments.push({{groupName: 'Unassigned', color: 'rgba(150,150,150,0.7)', items: unassignedItems, collapsed: false}});

  // Flatten to ordered x-axis labels (collapsed groups → group name; unassigned → sample names)
  var orderedNames = [];
  segments.forEach(function(seg) {{
    if (seg.collapsed) {{
      orderedNames.push(seg.groupName);
    }} else {{
      seg.items.forEach(function(s) {{ orderedNames.push(s); }});
    }}
  }});

  if (!orderedNames.length) {{
    Plotly.purge('sp-chart');
    return;
  }}

  // Build traces per group-segment
  var traces = [];
  var shapes = [];
  var xCursor = 0;

  segments.forEach(function(seg, si) {{
    var xVals = [];
    var yMeans = [];
    var ySDs = [];
    var barColors = [];
    var scatterX = [];
    var scatterY = [];
    var scatterColors = [];
    var scatterText = [];

    if (seg.collapsed) {{
      // One bar for the entire group; individual points are all sample values
      var allVals = [], anyFlagged = false, sampleMeans = [];
      seg.items.forEach(function(sname) {{
        var d = allData.find(function(x) {{ return x.name === sname; }});
        if (!d) return;
        if (d.anyFlagged) anyFlagged = true;
        var vals = spGetVals(d);
        sampleMeans.push(vals.mean);
        vals.values.forEach(function(v) {{
          allVals.push({{v: v, sname: sname, flagged: d.anyFlagged || false}});
        }});
      }});
      var grpMean = sampleMeans.length ? sampleMeans.reduce(function(a,b){{return a+b;}},0)/sampleMeans.length : 0;
      var grpSD = 0;
      if (allVals.length > 1) {{
        var vm = allVals.reduce(function(a,b){{return a+b.v;}},0)/allVals.length;
        grpSD = Math.sqrt(allVals.reduce(function(a,b){{return a+Math.pow(b.v-vm,2);}},0)/(allVals.length-1));
      }}
      xVals = [seg.groupName];
      yMeans = [grpMean];
      ySDs = [grpSD];
      barColors = [anyFlagged ? 'rgba(200,50,50,0.8)' : seg.color];
      allVals.forEach(function(pt) {{
        scatterX.push(seg.groupName);
        scatterY.push(pt.v);
        scatterColors.push(pt.flagged ? 'rgba(180,20,20,0.9)' : seg.color);
        scatterText.push(pt.sname);
      }});
    }} else {{
      // Individual bar per sample (unassigned pool)
      seg.items.forEach(function(sname) {{
        var d = allData.find(function(x) {{ return x.name === sname; }});
        if (!d) return;
        var flagged = d.anyFlagged;
        var vals = spGetVals(d);
        xVals.push(sname);
        yMeans.push(vals.mean);
        ySDs.push(vals.sd);
        barColors.push(flagged ? 'rgba(200,50,50,0.8)' : seg.color);
        vals.values.forEach(function(v) {{
          scatterX.push(sname);
          scatterY.push(v);
          scatterColors.push(flagged ? 'rgba(180,20,20,0.9)' : seg.color);
          scatterText.push(sname);
        }});
      }});
    }}

    // Separator shape before this segment (except the first)
    if (si > 0 && xCursor > 0) {{
      shapes.push({{
        type: 'line',
        xref: 'x', yref: 'paper',
        x0: xCursor - 0.5, x1: xCursor - 0.5,
        y0: 0, y1: 1,
        line: {{ color: '#aaa', width: 1, dash: 'dot' }}
      }});
    }}

    // Bar trace
    traces.push({{
      type: 'bar',
      name: seg.groupName,
      x: xVals,
      y: yMeans,
      error_y: {{
        type: 'data',
        array: ySDs,
        visible: true,
        color: '#444',
        thickness: 1.5,
        width: 4
      }},
      marker: {{ color: barColors }},
      showlegend: true,
      legendgroup: seg.groupName,
      hovertemplate: '<b>%{{x}}</b><br>Mean: %{{y:.4g}}<extra>' + seg.groupName + '</extra>'
    }});

    // Scatter trace for individual points
    if (scatterX.length) {{
      traces.push({{
        type: 'scatter',
        mode: 'markers',
        name: seg.groupName + ' pts',
        x: scatterX,
        y: scatterY,
        marker: {{
          color: scatterColors,
          size: 6,
          symbol: 'circle',
          line: {{ color: 'rgba(0,0,0,0.4)', width: 1 }}
        }},
        text: scatterText,
        showlegend: false,
        legendgroup: seg.groupName,
        hovertemplate: '<b>%{{text}}</b><br>Value: %{{y:.4g}}<extra></extra>'
      }});
    }}

    xCursor += seg.collapsed ? 1 : seg.items.length;
  }});

  var layout = {{
    barmode: 'group',
    height: 460,
    margin: {{ l: 100, r: 40, t: 40, b: 160 }},
    xaxis: {{
      tickangle: -40,
      automargin: true,
      categoryorder: 'array',
      categoryarray: orderedNames
    }},
    yaxis: {{
      title: {{ text: yTitle, standoff: 12 }},
      automargin: false,
      rangemode: 'tozero'
    }},
    shapes: shapes,
    legend: {{ orientation: 'h', x: 0, y: 1.08 }},
    paper_bgcolor: 'white',
    plot_bgcolor: 'white'
  }};

  Plotly.react('sp-chart', traces, layout, {{responsive: true}});
}}
// ── QC Plots Tab ─────────────────────────────────────────────────────────────
function qpNextColor() {{
  return QP_PALETTE[qpGroups.length % QP_PALETTE.length];
}}

function qpApplyDefaultGrouping() {{
  // Auto-group unassigned samples by QC level
  var levelSamples = {{}};
  (QP_DATA.levels || []).forEach(function(level) {{
    levelSamples[level] = [];
  }});
  var remaining = [];
  qpUnassigned.forEach(function(sname) {{
    var allSamples = QP_DATA.samples[qpCurrentAnalyte] || [];
    var entry = allSamples.find(function(s) {{ return s.name === sname; }});
    if (entry && entry.level && levelSamples.hasOwnProperty(entry.level)) {{
      levelSamples[entry.level].push(sname);
    }} else {{
      remaining.push(sname);
    }}
  }});
  qpUnassigned = remaining;
  (QP_DATA.levels || []).forEach(function(level) {{
    if (levelSamples[level].length > 0) {{
      qpGroups.push({{
        id: ++qpGroupIdCounter,
        name: level,
        color: QP_PALETTE[qpGroups.length % QP_PALETTE.length],
        visible: true,
        samples: levelSamples[level]
      }});
    }}
  }});
}}

function qpInit() {{
  if (!qpInitialized) {{
    qpInitialized = true;
    qpGroups = [];
    qpGroupIdCounter = 0;
    qpSortMode = 'group';
    var analytes = QP_DATA.analytes || [];
    qpCurrentAnalyte = analytes.length > 0 ? analytes[0] : null;
    qpUnassigned = qpCurrentAnalyte && QP_DATA.samples[qpCurrentAnalyte]
      ? QP_DATA.samples[qpCurrentAnalyte].map(function(s) {{ return s.name; }})
      : [];
    qpApplyDefaultGrouping();
    qpBuildAnalyteBar();
  }}
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpBuildAnalyteBar() {{
  var bar = document.getElementById('qp-analyte-bar');
  if (!bar) return;
  bar.innerHTML = '';
  (QP_DATA.analytes || []).forEach(function(a) {{
    var btn = document.createElement('button');
    btn.className = 'sp-analyte-btn' + (a === qpCurrentAnalyte ? ' active' : '');
    btn.textContent = a;
    btn.onclick = function() {{ qpSelectAnalyte(a); }};
    bar.appendChild(btn);
  }});
}}

function qpSelectAnalyte(name) {{
  qpCurrentAnalyte = name;
  qpGroups = [];
  qpGroupIdCounter = 0;
  var allSamples = QP_DATA.samples[name] ? QP_DATA.samples[name].map(function(s) {{ return s.name; }}) : [];
  qpUnassigned = allSamples.slice();
  qpApplyDefaultGrouping();
  document.querySelectorAll('#qp-analyte-bar .sp-analyte-btn').forEach(function(b) {{
    b.classList.toggle('active', b.textContent === name);
  }});
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpGetSampleData(sname) {{
  if (!qpCurrentAnalyte || !QP_DATA.samples[qpCurrentAnalyte]) return null;
  var arr = QP_DATA.samples[qpCurrentAnalyte];
  for (var i = 0; i < arr.length; i++) {{
    if (arr[i].name === sname) return arr[i];
  }}
  return null;
}}

function qpRenderGroupPanel() {{
  qpLastCheckedIdx = -1;
  var pool = document.getElementById('qp-unassigned-pool');
  if (pool) {{
    pool.innerHTML = '';
    qpUnassigned.forEach(function(sname) {{
      pool.appendChild(qpMakeChip(sname, '__unassigned__'));
    }});
  }}
  var gc = document.getElementById('qp-groups-container');
  if (!gc) return;
  gc.innerHTML = '';
  qpGroups.forEach(function(g) {{
    var block = document.createElement('div');
    block.className = 'sp-group-block';
    var hdr = document.createElement('div');
    hdr.className = 'sp-group-header';
    hdr.style.background = g.color;
    var isFirst = (qpGroups.indexOf(g) === 0);
    var isLast  = (qpGroups.indexOf(g) === qpGroups.length - 1);
    var btnStyle = 'background:rgba(255,255,255,0.25);color:white;border-color:rgba(255,255,255,0.4);';
    var btnDisabled = 'background:rgba(255,255,255,0.08);color:rgba(255,255,255,0.3);border-color:rgba(255,255,255,0.15);cursor:default;';
    hdr.innerHTML =
      '<span style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="' + g.name + '">' + g.name + '</span>' +
      '<button class="sp-btn sp-btn-icon" style="' + (isFirst ? btnDisabled : btnStyle) + '" ' +
        (isFirst ? 'disabled ' : 'onclick="qpMoveGroup(' + g.id + ',-1)" ') + 'title="Move up">&#x25B4;</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + (isLast  ? btnDisabled : btnStyle) + '" ' +
        (isLast  ? 'disabled ' : 'onclick="qpMoveGroup(' + g.id + ',1)" ')  + 'title="Move down">&#x25BE;</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + btnStyle + '" ' +
        'onclick="qpRenameGroup(' + g.id + ')" title="Rename group">&#x270F;</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + btnStyle + '" ' +
        'onclick="qpToggleGroup(' + g.id + ')" title="Show/hide in chart">' + (g.visible ? '&#128065;' : '&#128564;') + '</button>' +
      '<button class="sp-btn sp-btn-icon" style="' + btnStyle + '" ' +
        'onclick="qpDeleteGroup(' + g.id + ')" title="Delete group">&#x2715;</button>';
    block.appendChild(hdr);
    var dz = document.createElement('div');
    dz.className = 'sp-group-drop sp-drop-zone';
    dz.setAttribute('ondragover', 'qpDragOver(event)');
    dz.setAttribute('ondrop', "qpDrop(event,'" + g.id + "')");
    g.samples.forEach(function(sname) {{
      dz.appendChild(qpMakeChip(sname, g.id));
    }});
    block.appendChild(dz);
    gc.appendChild(block);
  }});
}}

function qpMakeChip(sname, groupId) {{
  var span = document.createElement('span');
  span.className = 'sp-chip';
  span.draggable = true;
  span.title = sname;
  var label = sname;
  if (groupId === '__unassigned__') {{
    var cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.setAttribute('data-sample', sname);
    cb.onclick = function(e) {{
      e.stopPropagation();
      var idx = qpUnassigned.indexOf(sname);
      var allCbs = Array.from(document.querySelectorAll('#qp-unassigned-pool input[type=checkbox]'));
      if (e.shiftKey && qpLastCheckedIdx >= 0 && idx !== qpLastCheckedIdx) {{
        var lo = Math.min(qpLastCheckedIdx, idx);
        var hi = Math.max(qpLastCheckedIdx, idx);
        var newState = allCbs[idx].checked;
        for (var i = lo; i <= hi; i++) {{
          if (allCbs[i]) allCbs[i].checked = newState;
        }}
      }}
      qpLastCheckedIdx = idx;
    }};
    span.appendChild(cb);
  }}
  var txt = document.createElement('span');
  txt.textContent = label;
  txt.style.maxWidth = '120px';
  txt.style.overflow = 'hidden';
  txt.style.textOverflow = 'ellipsis';
  txt.style.whiteSpace = 'nowrap';
  span.appendChild(txt);
  span.addEventListener('dragstart', function(e) {{ qpDragStart(e, sname, groupId); }});
  span.addEventListener('dragend', function(e) {{ qpDragEnd(e); }});
  return span;
}}

function qpCreateGroup() {{
  var name = prompt('Group name:');
  if (!name || !name.trim()) return;
  name = name.trim();
  qpGroups.push({{
    id: ++qpGroupIdCounter,
    name: name,
    color: qpNextColor(),
    visible: true,
    samples: []
  }});
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpSelectAllUnassigned(btn) {{
  var cbs = document.querySelectorAll('#qp-unassigned-pool input[type=checkbox]');
  var allChecked = Array.from(cbs).every(function(cb) {{ return cb.checked; }});
  cbs.forEach(function(cb) {{ cb.checked = !allChecked; }});
  btn.textContent = allChecked ? 'Select All' : 'Deselect All';
  qpLastCheckedIdx = -1;
}}

function qpAssignChecked() {{
  var checked = [];
  document.querySelectorAll('#qp-unassigned-pool input[type=checkbox]:checked').forEach(function(cb) {{
    checked.push(cb.getAttribute('data-sample'));
  }});
  if (!checked.length) {{ alert('Check at least one sample first.'); return; }}
  var name = prompt('Assign to group name:');
  if (!name || !name.trim()) return;
  name = name.trim();
  var g = qpGroups.find(function(x) {{ return x.name === name; }});
  if (!g) {{
    g = {{ id: ++qpGroupIdCounter, name: name, color: qpNextColor(), visible: true, samples: [] }};
    qpGroups.push(g);
  }}
  checked.forEach(function(s) {{
    qpUnassigned = qpUnassigned.filter(function(u) {{ return u !== s; }});
    if (g.samples.indexOf(s) === -1) g.samples.push(s);
  }});
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpDeleteGroup(id) {{
  var g = qpGroups.find(function(x) {{ return x.id == id; }});
  if (!g) return;
  g.samples.forEach(function(s) {{
    if (qpUnassigned.indexOf(s) === -1) qpUnassigned.push(s);
  }});
  qpGroups = qpGroups.filter(function(x) {{ return x.id != id; }});
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpToggleGroup(id) {{
  var g = qpGroups.find(function(x) {{ return x.id == id; }});
  if (!g) return;
  g.visible = !g.visible;
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpRenameGroup(id) {{
  var g = qpGroups.find(function(x) {{ return x.id == id; }});
  if (!g) return;
  var newName = prompt('Rename group:', g.name);
  if (newName === null || newName.trim() === '') return;
  g.name = newName.trim();
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpMoveGroup(id, dir) {{
  var idx = qpGroups.findIndex(function(x) {{ return x.id == id; }});
  var newIdx = idx + dir;
  if (newIdx < 0 || newIdx >= qpGroups.length) return;
  var tmp = qpGroups[idx];
  qpGroups[idx] = qpGroups[newIdx];
  qpGroups[newIdx] = tmp;
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpSetSort(mode, btn) {{
  qpSortMode = mode;
  document.querySelectorAll('#tab-qcplots .sp-sort-btn').forEach(function(b) {{ b.classList.remove('active-sort'); }});
  btn.classList.add('active-sort');
  qpRenderChart();
}}

function qpDragStart(e, name, fromGroup) {{
  qpDragPayload = {{name: name, fromGroup: fromGroup}};
  e.dataTransfer.effectAllowed = 'move';
  e.currentTarget.style.opacity = '0.4';
}}

function qpDragEnd(e) {{
  e.currentTarget.style.opacity = '1';
  document.querySelectorAll('#tab-qcplots .sp-drop-zone').forEach(function(z) {{
    z.classList.remove('drag-over');
  }});
}}

function qpDragOver(e) {{
  e.preventDefault();
  e.dataTransfer.dropEffect = 'move';
  var zone = e.currentTarget.closest('.sp-drop-zone');
  if (zone) zone.classList.add('drag-over');
}}

function qpDrop(e, toGroupId) {{
  e.preventDefault();
  document.querySelectorAll('#tab-qcplots .sp-drop-zone').forEach(function(z) {{ z.classList.remove('drag-over'); }});
  if (!qpDragPayload) return;
  var name = qpDragPayload.name;
  var fromGroup = qpDragPayload.fromGroup;
  qpDragPayload = null;
  if (fromGroup === toGroupId) return;
  if (fromGroup === '__unassigned__') {{
    qpUnassigned = qpUnassigned.filter(function(s) {{ return s !== name; }});
  }} else {{
    var sg = qpGroups.find(function(g) {{ return g.id == fromGroup; }});
    if (sg) sg.samples = sg.samples.filter(function(s) {{ return s !== name; }});
  }}
  if (toGroupId === '__unassigned__') {{
    if (qpUnassigned.indexOf(name) === -1) qpUnassigned.push(name);
  }} else {{
    var tg = qpGroups.find(function(g) {{ return g.id == toGroupId; }});
    if (tg && tg.samples.indexOf(name) === -1) tg.samples.push(name);
  }}
  qpRenderGroupPanel();
  qpRenderChart();
}}

function qpRenderChart() {{
  if (!qpCurrentAnalyte || !QP_DATA.samples[qpCurrentAnalyte]) {{
    Plotly.purge('qp-chart');
    return;
  }}
  var allData = QP_DATA.samples[qpCurrentAnalyte];
  var showUnassigned = document.getElementById('qp-show-unassigned') ? document.getElementById('qp-show-unassigned').checked : true;
  var units = QP_DATA.units || '';
  var yTitle = qpCurrentAnalyte + ' Concentration' + (units ? ' (' + units + ')' : '');

  var segments = [];

  qpGroups.forEach(function(g) {{
    if (!g.visible) return;
    var items = g.samples.filter(function(s) {{
      return allData.some(function(d) {{ return d.name === s; }});
    }});
    if (qpSortMode === 'asc') {{
      items.sort(function(a, b) {{
        var da = allData.find(function(d) {{ return d.name === a; }});
        var db = allData.find(function(d) {{ return d.name === b; }});
        return (da ? da.mean : 0) - (db ? db.mean : 0);
      }});
    }} else if (qpSortMode === 'desc') {{
      items.sort(function(a, b) {{
        var da = allData.find(function(d) {{ return d.name === a; }});
        var db = allData.find(function(d) {{ return d.name === b; }});
        return (db ? db.mean : 0) - (da ? da.mean : 0);
      }});
    }}
    if (items.length) segments.push({{groupName: g.name, color: g.color, items: items, collapsed: true}});
  }});

  var unassignedItems = showUnassigned
    ? qpUnassigned.filter(function(s) {{ return allData.some(function(d) {{ return d.name === s; }}); }})
    : [];
  if (qpSortMode === 'asc') {{
    unassignedItems.sort(function(a, b) {{
      var da = allData.find(function(d) {{ return d.name === a; }});
      var db = allData.find(function(d) {{ return d.name === b; }});
      return (da ? da.mean : 0) - (db ? db.mean : 0);
    }});
  }} else if (qpSortMode === 'desc') {{
    unassignedItems.sort(function(a, b) {{
      var da = allData.find(function(d) {{ return d.name === a; }});
      var db = allData.find(function(d) {{ return d.name === b; }});
      return (db ? db.mean : 0) - (da ? da.mean : 0);
    }});
  }}
  if (unassignedItems.length) segments.push({{groupName: 'Unassigned', color: 'rgba(150,150,150,0.7)', items: unassignedItems, collapsed: false}});

  var orderedNames = [];
  segments.forEach(function(seg) {{
    if (seg.collapsed) {{
      orderedNames.push(seg.groupName);
    }} else {{
      seg.items.forEach(function(s) {{ orderedNames.push(s); }});
    }}
  }});

  if (!orderedNames.length) {{
    Plotly.purge('qp-chart');
    return;
  }}

  var traces = [];
  var shapes = [];
  var xCursor = 0;

  segments.forEach(function(seg, si) {{
    var xVals = [];
    var yMeans = [];
    var ySDs = [];
    var barColors = [];
    var scatterX = [];
    var scatterY = [];
    var scatterColors = [];
    var scatterText = [];

    if (seg.collapsed) {{
      var allVals = [], sampleMeans = [];
      seg.items.forEach(function(sname) {{
        var d = allData.find(function(x) {{ return x.name === sname; }});
        if (!d) return;
        sampleMeans.push(d.mean);
        (d.values || []).forEach(function(v) {{
          allVals.push({{v: v, sname: sname}});
        }});
      }});
      var grpMean = sampleMeans.length ? sampleMeans.reduce(function(a,b){{return a+b;}},0)/sampleMeans.length : 0;
      var grpSD = 0;
      if (allVals.length > 1) {{
        var vm = allVals.reduce(function(a,b){{return a+b.v;}},0)/allVals.length;
        grpSD = Math.sqrt(allVals.reduce(function(a,b){{return a+Math.pow(b.v-vm,2);}},0)/(allVals.length-1));
      }}
      xVals = [seg.groupName];
      yMeans = [grpMean];
      ySDs = [grpSD];
      barColors = [seg.color];
      allVals.forEach(function(pt) {{
        scatterX.push(seg.groupName);
        scatterY.push(pt.v);
        scatterColors.push(seg.color);
        scatterText.push(pt.sname);
      }});
    }} else {{
      seg.items.forEach(function(sname) {{
        var d = allData.find(function(x) {{ return x.name === sname; }});
        if (!d) return;
        xVals.push(sname);
        yMeans.push(d.mean);
        ySDs.push(d.sd);
        barColors.push(seg.color);
        (d.values || []).forEach(function(v) {{
          scatterX.push(sname);
          scatterY.push(v);
          scatterColors.push(seg.color);
          scatterText.push(sname);
        }});
      }});
    }}

    if (si > 0 && xCursor > 0) {{
      shapes.push({{
        type: 'line',
        xref: 'x', yref: 'paper',
        x0: xCursor - 0.5, x1: xCursor - 0.5,
        y0: 0, y1: 1,
        line: {{ color: '#aaa', width: 1, dash: 'dot' }}
      }});
    }}

    traces.push({{
      type: 'bar',
      name: seg.groupName,
      x: xVals,
      y: yMeans,
      error_y: {{
        type: 'data',
        array: ySDs,
        visible: true,
        color: '#444',
        thickness: 1.5,
        width: 4
      }},
      marker: {{ color: barColors }},
      showlegend: true,
      legendgroup: seg.groupName,
      hovertemplate: '<b>%{{x}}</b><br>Mean: %{{y:.4g}}<extra>' + seg.groupName + '</extra>'
    }});

    if (scatterX.length) {{
      traces.push({{
        type: 'scatter',
        mode: 'markers',
        name: seg.groupName + ' pts',
        x: scatterX,
        y: scatterY,
        marker: {{
          color: scatterColors,
          size: 6,
          symbol: 'circle',
          line: {{ color: 'rgba(0,0,0,0.4)', width: 1 }}
        }},
        text: scatterText,
        showlegend: false,
        legendgroup: seg.groupName,
        hovertemplate: '<b>%{{text}}</b><br>Value: %{{y:.4g}}<extra></extra>'
      }});
    }}

    xCursor += seg.collapsed ? 1 : seg.items.length;
  }});

  // Expected concentration reference lines (±30% band)
  var expConc = QP_DATA.expected && QP_DATA.expected[qpCurrentAnalyte];
  if (expConc) {{
    shapes.push({{type:'rect', xref:'paper', yref:'y', x0:0, x1:1,
                 y0: expConc*0.7, y1: expConc*1.3,
                 fillcolor:'rgba(255,165,0,0.12)', line:{{width:0}}}});
    shapes.push({{type:'line', xref:'paper', yref:'y', x0:0, x1:1,
                 y0: expConc, y1: expConc,
                 line:{{color:'orange', width:1.5, dash:'dash'}}}});
  }}

  var layout = {{
    barmode: 'group',
    height: 460,
    margin: {{ l: 100, r: 40, t: 40, b: 160 }},
    xaxis: {{
      tickangle: -40,
      automargin: true,
      categoryorder: 'array',
      categoryarray: orderedNames
    }},
    yaxis: {{
      title: {{ text: yTitle, standoff: 12 }},
      automargin: false,
      rangemode: 'tozero'
    }},
    shapes: shapes,
    legend: {{ orientation: 'h', x: 0, y: 1.08 }},
    paper_bgcolor: 'white',
    plot_bgcolor: 'white'
  }};

  Plotly.react('qp-chart', traces, layout, {{responsive: true}});
}}
// ── End QC Plots Tab ──────────────────────────────────────────────────────────

// ── End Sample Plots Tab ─────────────────────────────────────────────────────
</script>
</body>
</html>"""

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"Saved HTML report: {html_path}")


def run_analysis(msd_path, platemap_path, output_path, spots_override=None, units=None, cv_threshold=25, dilution_factors=None, lloq_method='current', total_protein_path=None, qc_dilution_factors=None, qc_expected_concentrations=None, group_dilution_factors=None):
    _ensure_deps()   # lazy-load numpy / scipy / matplotlib / openpyxl
    print("=" * 60)
    print("MSD 4PL ANALYSIS")
    print("=" * 60)

    print(f"\nParsing MSD file: {msd_path}")
    plates = parse_msd_file(msd_path)
    for p in plates:
        n = spots_override if spots_override else p['spots_per_well']
        p['spots_per_well'] = n
        print(f"  Plate {p['plate_num']}: {len(p['data'])} wells x {n} spots")

    print(f"\nParsing plate map: {platemap_path}")
    plate_maps, raw_plate_blocks = parse_plate_map_grid(platemap_path)
    n_plate_maps = len(plate_maps)
    print(f"  Found {n_plate_maps} plate map(s)")

    if len(plates) != n_plate_maps:
        msg = (f"MSD file contains {len(plates)} plate(s), but plate map contains "
               f"{n_plate_maps} plate map(s). They must match.")
        print(f"Error: {msg}")
        raise RuntimeError(msg)

    try:
        plate_dilution_factors = parse_plate_dilution_factors(dilution_factors, len(plates))
    except ValueError as e:
        print(f"Error parsing dilution factors: {e}")
        raise

    # Build per-plate well lookups (well → list of entries; a well can serve multiple groups)
    plate_well_maps = {}
    for pm_num, entries in plate_maps.items():
        wm = {}
        for e in entries:
            wm.setdefault(normalize_well(e['well']), []).append(e)
        plate_well_maps[pm_num] = wm

    results = []
    for plate_data in plates:
        pnum = plate_data['plate_num']
        n_spots = plate_data['spots_per_well']
        wd = plate_data['data']

        # Match plate map: use matching plate number, or fall back to plate 1 if only one map
        if pnum in plate_well_maps:
            well_map = plate_well_maps[pnum]
        elif n_plate_maps == 1:
            well_map = list(plate_well_maps.values())[0]
            print(f"\n  (Using single plate map for MSD Plate {pnum})")
        else:
            print(f"\n  ⚠ No plate map found for MSD Plate {pnum} — skipping")
            continue

        for spot_idx in range(n_spots):
            spot_num = spot_idx + 1

            # Collect all well data for this spot, tagged with group.
            # A single well can produce multiple entries (multi-group & standard syntax).
            spot_wells = []
            for well_id, spot_signals in wd.items():
                nw = normalize_well(well_id)
                if spot_idx >= len(spot_signals):
                    continue
                signal = spot_signals[spot_idx]
                info_list = well_map.get(nw)
                if not info_list:
                    continue
                for info in info_list:
                    spot_wells.append({
                        'well': nw, 'signal': signal,
                        'sample_type': info['sample_type'],
                        'concentration': info.get('concentration', np.nan),
                        'sample_name': info.get('sample_name', ''),
                        'group': info.get('group', '_default')
                    })

            # Determine unique groups on this plate
            groups = sorted(set(w['group'] for w in spot_wells if w['group'] != '_default'))
            if not groups:
                groups = ['_default']

            for group in groups:
                group_label = group if group != '_default' else ''
                label = f"Plate {pnum}, Spot {spot_num}" + (f", Group {group}" if group_label else "")
                print(f"\n── {label} ──")

                # Partition: standards/unknowns/blanks for this group
                # Blanks with '_default' group are shared across all groups
                standards, unknowns, blanks = [], [], []
                for w in spot_wells:
                    wg = w['group']
                    stype = w['sample_type']

                    if stype == 'Blank' and (wg == group or wg == '_default'):
                        blanks.append({'well': w['well'], 'signal': w['signal'],
                                       'sample_name': w['sample_name']})
                    elif wg != group:
                        continue
                    elif stype == 'Standard':
                        conc = w['concentration']
                        if pd.notna(conc) and conc > 0:
                            standards.append({'well': w['well'], 'conc': conc, 'signal': w['signal']})
                    elif stype == 'Unknown':
                        unknowns.append({'well': w['well'], 'signal': w['signal'],
                                         'sample_name': w['sample_name']})

                no_standards = not bool(standards)
                if no_standards:
                    print("  ⚠ No standards detected for this curve")
                    params, r2 = None, None
                else:
                    conc_list = [s['conc'] for s in standards] + [0] * len(blanks)
                    signal_list = [s['signal'] for s in standards] + [b['signal'] for b in blanks]
                    params, r2 = fit_4pl(conc_list, signal_list)

                if params is not None:
                    a, b, c, d = params
                    print(f"  a={a:.1f}  b={b:.4f}  c={c:.2f}  d={d:.1f}  R²={r2:.6f}")
                    for u in unknowns:
                        try:
                            u['interp_conc'] = inverse_4pl(u['signal'], *params)
                        except (ValueError, ZeroDivisionError, OverflowError):
                            u['interp_conc'] = np.nan
                else:
                    if not no_standards:
                        print("  ⚠ Curve fit FAILED")
                    for u in unknowns:
                        u['interp_conc'] = np.nan

                unknowns.sort(key=lambda x: (x['well'][0], int(re.search(r'\d+', x['well']).group())))

                blank_sigs = [b['signal'] for b in blanks if np.isfinite(b['signal'])]
                lloq_sig_cached = calculate_lloq_signal(blank_sigs, lloq_method)
                results.append({
                    'plate': pnum, 'spot': spot_num, 'group': group_label,
                    'params': params, 'r2': r2,
                    'standards': sorted(standards, key=lambda x: x['conc']),
                    'unknowns': unknowns, 'blanks': blanks,
                    'no_standards': no_standards,
                    'lloq_sig': lloq_sig_cached
                })

    missing = [r for r in results if r.get('no_standards')]
    if missing:
        labels = []
        for r in missing:
            lbl = f"Plate {r['plate']}, Spot {r['spot']}"
            if r.get('group'):
                lbl += f", Group {r['group']}"
            labels.append(lbl)
        msg = "Standards are missing for: " + "; ".join(labels)
        print(f"Error: {msg}")
        raise RuntimeError(msg)

    # Parse total protein CSV if provided
    total_protein_map = None
    if total_protein_path:
        try:
            total_protein_map = parse_total_protein_csv(total_protein_path)
            print(f"\nLoaded total protein data: {len(total_protein_map)} animal/tissue entries")
        except Exception as e:
            print(f"Warning: could not load total protein CSV: {e}")

    print(f"\n{'=' * 60}")
    print(f"Generating Excel: {output_path}")
    create_output(results, output_path, msd_path, raw_plate_blocks, units, cv_threshold, plate_dilution_factors, lloq_method, total_protein_map, qc_dilution_factors, qc_expected_concentrations, group_dilution_factors=group_dilution_factors)
    print("Done!")

    # Generate and open interactive HTML report (Excel is opened from within HTML)
    # Write to system temp dir to avoid macOS Desktop-folder TCC permission prompt.
    # The "Open Excel" link inside the HTML still uses the full Desktop path,
    # which the browser can follow without involving this app.
    import tempfile
    html_basename = os.path.splitext(os.path.basename(output_path))[0] + '.html'
    html_path = os.path.join(tempfile.gettempdir(), html_basename)
    try:
        generate_html_report(results, html_path, msd_path, units,
                             qc_dilution_factors, qc_expected_concentrations,
                             plate_dilution_factors, lloq_method,
                             total_protein_map, output_path,
                             group_dilution_factors=group_dilution_factors,
                             cv_threshold=cv_threshold)
        if os.path.exists(html_path):
            _open_file(html_path)
    except Exception as e:
        import traceback as _tb
        print(f"Warning: HTML report could not be generated: {e}")
        _tb.print_exc()

    # Save last run parameters
    last_args = {
        'msd': msd_path,
        'platemap': platemap_path,
        'output': output_path,
        'spots': spots_override,
        'units': units,
        'cv_threshold': cv_threshold,
        'dilution_factors': list(plate_dilution_factors.values()) if plate_dilution_factors else None,
        'lloq_method': lloq_method,
        'total_protein': total_protein_path,
        'qc_dilution_factors': qc_dilution_factors,
        'qc_expected_concentrations': qc_expected_concentrations,
        'group_dilution_factors': group_dilution_factors,
        'status': 'pass',
    }
    _save_run_to_history(last_args)


def run_interactive():
    """Launch a single-page GUI for configuring all analysis options."""
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import json
    import os

    def browse_file(var, title, filetypes):
        filename = filedialog.askopenfilename(title=title, filetypes=filetypes)
        if filename:
            var.set(filename)

    def browse_save(var, title, defaultextension, filetypes, initialfile):
        filename = filedialog.asksaveasfilename(title=title, defaultextension=defaultextension, filetypes=filetypes, initialfile=initialfile)
        if filename:
            var.set(filename)

    def _apply_run_entry(entry):
        """Populate all form fields from a history entry dict."""
        msd_var.set(entry.get('msd', ''))
        platemap_var.set(entry.get('platemap', ''))
        output_var.set(entry.get('output', 'msd_4pl_results.xlsx'))
        spots_var.set(str(entry.get('spots') or ''))
        units_var.set(entry.get('units') or '')
        cv_threshold_var.set(str(entry.get('cv_threshold') or '25'))
        lloq_method_var.set(entry.get('lloq_method') or 'current')
        dilution_factors_var.set(
            ','.join(str(x) for x in entry['dilution_factors'])
            if entry.get('dilution_factors') else '')
        total_protein_var.set(entry.get('total_protein') or '')
        # Restore group dilution factors and per-group QC values if any were saved
        saved_grp = entry.get('group_dilution_factors') or {}
        saved_qc = entry.get('qc_dilution_factors') or {}
        saved_exp = entry.get('qc_expected_concentrations') or {}
        # Normalize old flat qc_dilution_factors format (skip silently)
        if saved_qc and not isinstance(next(iter(saved_qc.values()), {}), dict):
            saved_qc = {}
        if saved_exp and not isinstance(saved_exp, dict):
            saved_exp = {}
        if saved_grp:
            # Rebuild group rows for the saved groups
            for w in grp_rows_frame.winfo_children():
                w.destroy()
            group_df_vars.clear()
            grp_qc_vars.clear()
            grp_exp_vars.clear()
            # Collect all QC levels that appear in saved_qc across all groups
            all_saved_qc_cols = [lvl for lvl in QC_LEVELS
                                 if any(lvl in (saved_qc.get(g) or {}) for g in saved_grp)]
            col = 0
            ttk.Label(grp_rows_frame, text='Group', font=('TkDefaultFont', 9, 'bold')).grid(
                row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
            col += 1
            ttk.Label(grp_rows_frame, text='Dil. Factor', font=('TkDefaultFont', 9, 'bold')).grid(
                row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
            col += 1
            for lvl in all_saved_qc_cols:
                ttk.Label(grp_rows_frame, text=lvl, font=('TkDefaultFont', 9, 'bold')).grid(
                    row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
                col += 1
            ttk.Label(grp_rows_frame, text='Expected Conc.', font=('TkDefaultFont', 9, 'bold')).grid(
                row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
            for ri, (gname, gval) in enumerate(sorted(saved_grp.items()), 1):
                col = 0
                ttk.Label(grp_rows_frame, text=gname).grid(
                    row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
                col += 1
                df_var = tk.StringVar(value=str(gval))
                group_df_vars[gname] = df_var
                ttk.Entry(grp_rows_frame, textvariable=df_var, width=9).grid(
                    row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
                col += 1
                grp_qc_vars[gname] = {}
                g_qc = saved_qc.get(gname) or {}
                for lvl in all_saved_qc_cols:
                    if lvl in g_qc:
                        qc_var = tk.StringVar(value=str(g_qc[lvl]))
                        grp_qc_vars[gname][lvl] = qc_var
                        ttk.Entry(grp_rows_frame, textvariable=qc_var, width=8).grid(
                            row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
                    else:
                        ttk.Label(grp_rows_frame, text='—', foreground='grey').grid(
                            row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
                    col += 1
                exp_val = saved_exp.get(gname, '')
                exp_var = tk.StringVar(value=str(exp_val) if exp_val else '')
                grp_exp_vars[gname] = exp_var
                ttk.Entry(grp_rows_frame, textvariable=exp_var, width=10).grid(
                    row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
            grp_hint.config(text=f'Restored {len(saved_grp)} group(s) from history.')

    def load_selected_run():
        sel = history_lb.curselection()
        if not sel:
            messagebox.showinfo("No Selection", "Please click a run in the list to select it.")
            return
        idx = sel[0]
        history = _load_run_history()
        if idx >= len(history):
            return
        _apply_run_entry(history[idx])

    def _show_loading_screen():
        """Animated loading window with a 4PL sigmoid being drawn in real time."""
        import math, random, threading as _threading
        win = tk.Toplevel(root)
        win.title("MSD 4PL Analysis")
        win.resizable(False, False)
        win.configure(bg='white')
        win.protocol("WM_DELETE_WINDOW", lambda: None)   # prevent accidental close

        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        ww, wh = 460, 310
        win.geometry(f"{ww}x{wh}+{(sw-ww)//2}+{(sh-wh)//2}")
        win.grab_set()

        tk.Label(win, text="Running Analysis…", font=('Arial', 15, 'bold'),
                 bg='white', fg='#2c3e50').pack(pady=(20, 2))

        # ── Canvas ──────────────────────────────────────────────────────
        cw, ch = 420, 170
        canvas = tk.Canvas(win, width=cw, height=ch, bg='white', highlightthickness=0)
        canvas.pack(padx=20)

        # ── Status label & progress bar ─────────────────────────────────
        status_var = tk.StringVar(value="Parsing MSD data…")
        tk.Label(win, textvariable=status_var, font=('Arial', 10),
                 bg='white', fg='#555555').pack(pady=(6, 2))
        pb = ttk.Progressbar(win, mode='indeterminate', length=420)
        pb.pack(padx=20, pady=(0, 20))
        pb.start(12)

        # ── 4PL model for animation ──────────────────────────────────────
        def _4pl_anim(x):
            return 90000 + (200 - 90000) / (1 + (x / 10) ** 1.5)

        ml, mr, mt, mb = 48, 12, 14, 32          # margins
        pw, ph = cw - ml - mr, ch - mt - mb      # plot area
        xl, xh = math.log10(0.04), math.log10(600)
        yl, yh = math.log10(140),  math.log10(110000)

        def _px(x, y):
            px = ml + (math.log10(max(x, 1e-9)) - xl) / (xh - xl) * pw
            py = mt + ph - (math.log10(max(y, 1e-9)) - yl) / (yh - yl) * ph
            return px, py

        # Static axes
        canvas.create_line(ml, mt, ml, mt + ph, fill='#aaaaaa', width=1.5)
        canvas.create_line(ml, mt + ph, ml + pw, mt + ph, fill='#aaaaaa', width=1.5)
        canvas.create_text(ml + pw // 2, ch - 6, text='Concentration (log scale)',
                           font=('Arial', 7), fill='#888888')
        canvas.create_text(10, mt + ph // 2, text='Signal', angle=90,
                           font=('Arial', 7), fill='#888888')
        for lx in [math.log10(v) for v in [0.1, 1, 10, 100]]:
            px = ml + (lx - xl) / (xh - xl) * pw
            canvas.create_line(px, mt, px, mt + ph, fill='#eeeeee', width=1)

        # Curve smooth points
        _xs = [0.04 * (600 / 0.04) ** (i / 119) for i in range(120)]
        _curve_pts = [_px(x, _4pl_anim(x)) for x in _xs]

        # Scatter data (2 reps per conc, slight noise)
        random.seed(7)
        _concs = [0.1, 0.3, 1, 3, 10, 30, 100, 300]
        _scatter = []
        for c in _concs:
            for _ in range(2):
                _scatter.append(_px(c, _4pl_anim(c) * random.uniform(0.91, 1.09)))

        # Pre-create dot items (hidden)
        _dot_ids = []
        for px, py in _scatter:
            did = canvas.create_oval(px - 4, py - 4, px + 4, py + 4,
                                     fill='#2F5496', outline='#1a2f6e',
                                     width=1.2, state='hidden')
            _dot_ids.append(did)

        # Animation state
        _st = {'phase': 0, 'step': 0, 'after_id': None}

        def _animate():
            p, s = _st['phase'], _st['step']
            if p == 0:                          # dots pop in one by one
                if s < len(_dot_ids):
                    canvas.itemconfig(_dot_ids[s], state='normal')
                    _st['step'] += 1
                    _st['after_id'] = win.after(91, _animate)
                else:
                    _st['phase'], _st['step'] = 1, 0
                    _st['after_id'] = win.after(156, _animate)
            elif p == 1:                        # curve draws left → right
                if s < len(_curve_pts) - 2:
                    x1, y1 = _curve_pts[s]
                    x2, y2 = _curve_pts[s + 2]
                    canvas.create_line(x1, y1, x2, y2,
                                       fill='#E06C4A', width=2, tags='crv')
                    _st['step'] += 2
                    _st['after_id'] = win.after(26, _animate)
                else:
                    _st['phase'], _st['step'] = 2, 0
                    _st['after_id'] = win.after(26, _animate)
            elif p == 2:                        # hold
                _st['step'] += 1
                if _st['step'] > 35:
                    _st['phase'], _st['step'] = 3, 0
                _st['after_id'] = win.after(52, _animate)
            else:                               # reset
                canvas.delete('crv')
                for did in _dot_ids:
                    canvas.itemconfig(did, state='hidden')
                _st['phase'], _st['step'] = 0, 0
                _st['after_id'] = win.after(104, _animate)

        _animate()

        # Status message cycling
        _msgs = ["Parsing MSD data…", "Building plate maps…",
                 "Fitting 4PL curves…", "Calculating LLOQ values…",
                 "Writing Excel report…", "Generating HTML charts…"]
        _mi = [0]
        def _cycle():
            _mi[0] = (_mi[0] + 1) % len(_msgs)
            status_var.set(_msgs[_mi[0]])
            win._msg_id = win.after(2200, _cycle)
        win._msg_id = win.after(2200, _cycle)

        def _close():
            if _st['after_id']:
                win.after_cancel(_st['after_id'])
            if hasattr(win, '_msg_id'):
                win.after_cancel(win._msg_id)
            pb.stop()
            try:
                win.grab_release()
                win.destroy()
            except Exception:
                pass

        win.close_loading = _close
        return win

    def run():
        import threading as _threading
        import traceback as _traceback

        msd_path = msd_var.get().strip()
        platemap_path = platemap_var.get().strip()
        output_path = output_var.get().strip()
        spots_override = spots_var.get().strip()
        units = units_var.get().strip()
        cv_threshold = cv_threshold_var.get().strip()
        lloq_method = lloq_method_var.get()
        dilution_factors = dilution_factors_var.get().strip()
        total_protein_path = total_protein_var.get().strip()

        if not msd_path or not platemap_path or not output_path:
            messagebox.showerror("Error", "Please select MSD file, plate map, and output location.")
            return

        spots_override = int(spots_override) if spots_override and spots_override in ('1', '4', '10') else None
        units = units if units else None
        cv_threshold = float(cv_threshold) if cv_threshold else None
        dilution_factors = dilution_factors if dilution_factors else None
        total_protein_path = total_protein_path if total_protein_path else None

        # Collect group dilution factors
        group_dilution_factors = {}
        for gname, gvar in group_df_vars.items():
            val_str = gvar.get().strip()
            if val_str:
                try:
                    group_dilution_factors[gname] = float(val_str)
                except ValueError:
                    messagebox.showerror("Error", f"Invalid group dilution factor for '{gname}': '{val_str}'")
                    return
        group_dilution_factors = group_dilution_factors if group_dilution_factors else None

        # Collect per-group QC dilution factors
        qc_dilution_factors = {}
        for gname, level_vars in grp_qc_vars.items():
            for level, var in level_vars.items():
                val_str = var.get().strip()
                if val_str:
                    try:
                        qc_dilution_factors.setdefault(gname, {})[level] = float(val_str)
                    except ValueError:
                        messagebox.showerror("Error", f"Invalid QC dilution factor for {gname}/{level}: '{val_str}'")
                        return
        qc_dilution_factors = qc_dilution_factors if qc_dilution_factors else None

        # Collect per-group expected concentrations
        qc_expected_concentrations = {}
        for gname, var in grp_exp_vars.items():
            val_str = var.get().strip()
            if val_str:
                try:
                    qc_expected_concentrations[gname] = float(val_str)
                except ValueError:
                    messagebox.showerror("Error", f"Invalid expected concentration for {gname}: '{val_str}'")
                    return
        qc_expected_concentrations = qc_expected_concentrations if qc_expected_concentrations else None

        print(f"\nMSD file:   {msd_path}")
        print(f"Plate map:  {platemap_path}")
        print(f"Output:     {output_path}")
        if spots_override:
            print(f"Spots:      {spots_override} (manual override)")
        else:
            print("Spots:      auto-detect")
        if units:
            print(f"Units:      {units}")
        if cv_threshold is not None:
            print(f"CV threshold: {cv_threshold}")
        if dilution_factors:
            print(f"Dilution factors: {dilution_factors}")
        if group_dilution_factors:
            print(f"Group dilution factors: {group_dilution_factors}")
        print(f"LLOQ method: {lloq_method}")

        # Thread result container
        _result = {'error': None, 'done': False}

        def _worker():
            try:
                run_analysis(msd_path, platemap_path, output_path, spots_override,
                             units, cv_threshold, dilution_factors, lloq_method,
                             total_protein_path, qc_dilution_factors, qc_expected_concentrations,
                             group_dilution_factors=group_dilution_factors)
            except Exception as exc:
                _result['error'] = exc
                print(f"\nAnalysis error: {exc}")
                _traceback.print_exc()
            finally:
                _result['done'] = True

        loading = _show_loading_screen()
        _threading.Thread(target=_worker, daemon=True).start()

        def _poll():
            if not _result['done']:
                root.after(200, _poll)
                return
            loading.close_loading()
            if _result['error'] is not None:
                err_msg = str(_result['error']) or type(_result['error']).__name__
                fail_entry = {
                    'msd': msd_path, 'platemap': platemap_path,
                    'output': output_path, 'spots': spots_override,
                    'units': units, 'cv_threshold': cv_threshold,
                    'dilution_factors': dilution_factors,
                    'lloq_method': lloq_method,
                    'total_protein': total_protein_path,
                    'qc_dilution_factors': qc_dilution_factors,
                    'qc_expected_concentrations': qc_expected_concentrations,
                    'group_dilution_factors': group_dilution_factors,
                    'status': 'fail', 'error': err_msg,
                }
                _save_run_to_history(fail_entry)
                messagebox.showerror("Analysis Error", err_msg, parent=root)
                root.deiconify()
                refresh_history()
            else:
                root.destroy()

        root.withdraw()
        root.after(200, _poll)

    # ── Window setup ───────────────────────────────────────────────────
    root = tk.Tk()
    root.title("MSD 4PL Analysis Tool")
    root.geometry("860x720")
    root.minsize(760, 620)
    root.resizable(True, True)

    # ── Variables ──────────────────────────────────────────────────────
    msd_var = tk.StringVar()
    platemap_var = tk.StringVar()
    output_var = tk.StringVar(value="msd_4pl_results.xlsx")
    spots_var = tk.StringVar()
    units_var = tk.StringVar()
    cv_threshold_var = tk.StringVar(value="25")
    lloq_method_var = tk.StringVar(value="current")
    dilution_factors_var = tk.StringVar()
    total_protein_var = tk.StringVar()
    group_df_vars = {}   # {group: StringVar} — populated by _detect_groups
    grp_qc_vars = {}     # {group: {level: StringVar}} — populated by _detect_groups
    grp_exp_vars = {}    # {group: StringVar} — populated by _detect_groups

    # ── Header banner ──────────────────────────────────────────────────
    SLATE  = '#3a506b'   # soft slate-blue — clean, not corporate-heavy
    SLATE_LIGHT = '#c8d8e8'
    header_canvas = tk.Canvas(root, height=58, bg=SLATE, highlightthickness=0)
    header_canvas.pack(fill=tk.X, side=tk.TOP)
    header_canvas.create_text(18, 20, anchor='w', text='MSD 4PL Analysis Tool',
                              fill='white', font=('Helvetica', 16, 'bold'))
    header_canvas.create_text(18, 42, anchor='w',
                              text='4-Parameter Logistic Curve Fitting  ·  Quantitative Analysis',
                              fill=SLATE_LIGHT, font=('Helvetica', 10))

    # Thin accent rule below header
    tk.Canvas(root, height=2, bg='#7ba7bc', highlightthickness=0).pack(fill=tk.X)

    # ── Fixed bottom action bar (always visible, packed before scroll area) ──
    _bottom = ttk.Frame(root, padding='6 4 12 8')
    _bottom.pack(side=tk.BOTTOM, fill=tk.X)
    ttk.Separator(_bottom, orient='horizontal').pack(fill=tk.X, pady=(0, 8))
    _btn_row = ttk.Frame(_bottom)
    _btn_row.pack(fill=tk.X)
    # Buttons are added to _btn_row after `run` is defined (see bottom of this function)

    # ── Scrollable content area ────────────────────────────────────────
    _sc_host = ttk.Frame(root)
    _sc_host.pack(fill=tk.BOTH, expand=True)

    _vscroll = ttk.Scrollbar(_sc_host, orient=tk.VERTICAL)
    _vscroll.pack(side=tk.RIGHT, fill=tk.Y)

    _scroll_canvas = tk.Canvas(_sc_host, highlightthickness=0,
                                yscrollcommand=_vscroll.set)
    _scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    _vscroll.configure(command=_scroll_canvas.yview)

    outer = ttk.Frame(_scroll_canvas, padding='12 10 12 6')
    _cw = _scroll_canvas.create_window((0, 0), window=outer, anchor='nw')

    def _update_scrollregion(event=None):
        _scroll_canvas.configure(scrollregion=_scroll_canvas.bbox('all'))

    def _fit_canvas_width(event):
        _scroll_canvas.itemconfig(_cw, width=event.width)

    outer.bind('<Configure>', _update_scrollregion)
    _scroll_canvas.bind('<Configure>', _fit_canvas_width)

    def _on_mousewheel(event):
        if event.delta:
            _scroll_canvas.yview_scroll(int(-1 * event.delta / 120), 'units')
    _scroll_canvas.bind_all('<MouseWheel>', _on_mousewheel)

    outer.columnconfigure(0, weight=1)

    # helper: consistent row padding inside LabelFrames
    _rp = {'pady': 4}

    # ── Input Files ────────────────────────────────────────────────────
    files_lf = ttk.LabelFrame(outer, text='Input Files', padding='10 6')
    files_lf.pack(fill=tk.X, pady=(0, 8))
    files_lf.columnconfigure(1, weight=1)

    def _file_row(parent, row, label, var, btn_cmd, btn2_cmd=None):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky=tk.W,
                                           padx=(0, 10), **_rp)
        ttk.Entry(parent, textvariable=var, width=54).grid(row=row, column=1,
                                                            sticky=tk.EW, **_rp)
        ttk.Button(parent, text='Browse…', command=btn_cmd, width=8).grid(
            row=row, column=2, padx=(6, 0), **_rp)

    _file_row(files_lf, 0, 'MSD Data File:', msd_var,
              lambda: browse_file(msd_var, 'Select MSD Data File',
                                  [('MSD Text Files', '*.txt'), ('All Files', '*.*')]))
    _file_row(files_lf, 1, 'Plate Map CSV:', platemap_var,
              lambda: browse_file(platemap_var, 'Select Plate Map CSV',
                                  [('CSV Files', '*.csv'), ('All Files', '*.*')]))
    _file_row(files_lf, 2, 'Output Excel:', output_var,
              lambda: browse_save(output_var, 'Save Results As', '.xlsx',
                                  [('Excel Files', '*.xlsx')], 'msd_4pl_results.xlsx'))

    # ── Analysis Options ───────────────────────────────────────────────
    opts_lf = ttk.LabelFrame(outer, text='Analysis Options', padding='10 6')
    opts_lf.pack(fill=tk.X, pady=(0, 8))

    # Row 0 — Spots | Units
    ttk.Label(opts_lf, text='Spots per Well:').grid(row=0, column=0, sticky=tk.W, **_rp)
    spots_e = ttk.Entry(opts_lf, textvariable=spots_var, width=8)
    spots_e.grid(row=0, column=1, sticky=tk.W, padx=(0, 20), **_rp)
    ttk.Label(opts_lf, text='1, 4, 10 or blank', foreground='grey').grid(
        row=0, column=2, sticky=tk.W, padx=(0, 30), **_rp)
    ttk.Label(opts_lf, text='Units:').grid(row=0, column=3, sticky=tk.W, **_rp)
    ttk.Entry(opts_lf, textvariable=units_var, width=14).grid(
        row=0, column=4, sticky=tk.W, **_rp)
    ttk.Label(opts_lf, text='e.g. pg/mL', foreground='grey').grid(
        row=0, column=5, sticky=tk.W, padx=(4, 0), **_rp)

    # Row 1 — %CV | Dilution Factors
    ttk.Label(opts_lf, text='%CV Threshold:').grid(row=1, column=0, sticky=tk.W, **_rp)
    ttk.Entry(opts_lf, textvariable=cv_threshold_var, width=8).grid(
        row=1, column=1, sticky=tk.W, padx=(0, 20), **_rp)
    ttk.Label(opts_lf, text='Plate Dilution Factors:').grid(
        row=1, column=3, sticky=tk.W, **_rp)
    ttk.Entry(opts_lf, textvariable=dilution_factors_var, width=22).grid(
        row=1, column=4, columnspan=2, sticky=tk.W, **_rp)

    # Row 2 — LLOQ Method
    ttk.Label(opts_lf, text='LLOQ Method:').grid(row=2, column=0, sticky=tk.W, **_rp)
    lloq_inner = ttk.Frame(opts_lf)
    lloq_inner.grid(row=2, column=1, columnspan=5, sticky=tk.W, **_rp)
    ttk.Radiobutton(lloq_inner, text='Mean + 10×SD (current)',
                    variable=lloq_method_var, value='current').pack(side=tk.LEFT)
    ttk.Radiobutton(lloq_inner, text='3× Blank Mean',
                    variable=lloq_method_var, value='3xblank').pack(side=tk.LEFT, padx=(16, 0))

    # Row 3 — Total Protein (spans full width)
    ttk.Label(opts_lf, text='Total Protein CSV:').grid(row=3, column=0, sticky=tk.W, **_rp)
    ttk.Entry(opts_lf, textvariable=total_protein_var, width=46).grid(
        row=3, column=1, columnspan=4, sticky=tk.W, **_rp)
    ttk.Button(opts_lf, text='Browse…', width=8,
               command=lambda: browse_file(
                   total_protein_var, 'Select Total Protein CSV',
                   [('CSV Files', '*.csv'), ('All Files', '*.*')])).grid(
        row=3, column=5, padx=(6, 0), **_rp)

    # ── Group Dilution Factors ─────────────────────────────────────────
    grp_lf = ttk.LabelFrame(outer, text='Group Dilution Factors  (optional — applied per group detected in plate map)',
                             padding='10 6')
    grp_lf.pack(fill=tk.X, pady=(0, 8))

    grp_hint = ttk.Label(grp_lf,
                         text='Load a plate map, then click Detect Groups to set per-group dilution factors.',
                         foreground='grey')
    grp_hint.grid(row=0, column=0, columnspan=6, sticky=tk.W, pady=(0, 4))

    grp_btn_frame = ttk.Frame(grp_lf)
    grp_btn_frame.grid(row=1, column=0, columnspan=6, sticky=tk.W, pady=(0, 4))

    # Inner frame that holds the dynamically created group rows
    grp_rows_frame = ttk.Frame(grp_lf)
    grp_rows_frame.grid(row=2, column=0, columnspan=6, sticky=tk.EW)

    def _detect_groups():
        """Parse the platemap and create one dilution-factor row per group."""
        pm_path = platemap_var.get().strip()
        if not pm_path or not os.path.exists(pm_path):
            messagebox.showwarning("No Plate Map", "Please select a valid plate map CSV first.")
            return
        try:
            _ensure_deps()   # pandas / numpy needed by the parser
            plate_maps, _ = parse_plate_map_grid(pm_path)
        except Exception as exc:
            messagebox.showerror("Parse Error", f"Could not read plate map:\n{exc}")
            return

        # Collect all unique non-default groups across all plate maps
        found = set()
        for entries in plate_maps.values():
            for e in entries:
                g = e.get('group', '_default')
                if g and g != '_default':
                    found.add(g)

        # Destroy old rows
        for w in grp_rows_frame.winfo_children():
            w.destroy()

        if not found:
            ttk.Label(grp_rows_frame,
                      text='No named groups found in plate map (group prefix syntax: GroupName:value).',
                      foreground='grey').grid(row=0, column=0, columnspan=6, sticky=tk.W)
            group_df_vars.clear()
            return

        # Find QC levels per group
        grp_qc_levels = {g: set() for g in found}
        for entries in plate_maps.values():
            for e in entries:
                g = e.get('group', '_default')
                if g not in grp_qc_levels:
                    continue
                level = _identify_qc_level(e.get('sample_name', ''))
                if level:
                    grp_qc_levels[g].add(level)

        # All QC levels seen across all groups (for column headers)
        all_qc_cols = [lvl for lvl in QC_LEVELS if any(lvl in grp_qc_levels[g] for g in found)]

        # Preserve any existing values when re-detecting
        prev = {g: group_df_vars[g].get() for g in group_df_vars if g in found}
        prev_qc = {g: {lvl: grp_qc_vars[g][lvl].get() for lvl in grp_qc_vars[g]}
                   for g in grp_qc_vars if g in found}
        prev_exp = {g: grp_exp_vars[g].get() for g in grp_exp_vars if g in found}
        group_df_vars.clear()
        grp_qc_vars.clear()
        grp_exp_vars.clear()

        # Column headers: Group | Dil. Factor | [QC levels...] | Expected Conc.
        col = 0
        ttk.Label(grp_rows_frame, text='Group', font=('TkDefaultFont', 9, 'bold')).grid(
            row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
        col += 1
        ttk.Label(grp_rows_frame, text='Dil. Factor', font=('TkDefaultFont', 9, 'bold')).grid(
            row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
        col += 1
        for lvl in all_qc_cols:
            ttk.Label(grp_rows_frame, text=lvl, font=('TkDefaultFont', 9, 'bold')).grid(
                row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))
            col += 1
        ttk.Label(grp_rows_frame, text='Expected Conc.', font=('TkDefaultFont', 9, 'bold')).grid(
            row=0, column=col, sticky=tk.W, padx=(0, 8), pady=(0, 2))

        for ri, gname in enumerate(sorted(found), 1):
            col = 0
            ttk.Label(grp_rows_frame, text=gname).grid(
                row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
            col += 1
            df_var = tk.StringVar(value=prev.get(gname, ''))
            group_df_vars[gname] = df_var
            ttk.Entry(grp_rows_frame, textvariable=df_var, width=9).grid(
                row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
            col += 1
            grp_qc_vars[gname] = {}
            for lvl in all_qc_cols:
                if lvl in grp_qc_levels[gname]:
                    qc_var = tk.StringVar(value=prev_qc.get(gname, {}).get(lvl, ''))
                    grp_qc_vars[gname][lvl] = qc_var
                    ttk.Entry(grp_rows_frame, textvariable=qc_var, width=8).grid(
                        row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
                else:
                    ttk.Label(grp_rows_frame, text='—', foreground='grey').grid(
                        row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)
                col += 1
            exp_var = tk.StringVar(value=prev_exp.get(gname, ''))
            grp_exp_vars[gname] = exp_var
            ttk.Entry(grp_rows_frame, textvariable=exp_var, width=10).grid(
                row=ri, column=col, sticky=tk.W, padx=(0, 8), pady=2)

        grp_hint.config(text=f'Found {len(found)} group(s). Enter dilution factors and QC values per group (leave blank = 1× / no QC).')

    ttk.Button(grp_btn_frame, text='Detect Groups from Plate Map',
               command=_detect_groups).pack(side=tk.LEFT)
    ttk.Label(grp_btn_frame, text='  Priority: group factor > plate factor',
              foreground='grey').pack(side=tk.LEFT)

    # ── Previous Runs ──────────────────────────────────────────────────
    hist_lf = ttk.LabelFrame(outer, text='Previous Runs', padding='10 6')
    hist_lf.pack(fill=tk.X, pady=(0, 8))
    hist_lf.columnconfigure(0, weight=1)

    lb_outer = ttk.Frame(hist_lf)
    lb_outer.grid(row=0, column=0, sticky=tk.NSEW)
    lb_outer.columnconfigure(0, weight=1)
    lb_outer.rowconfigure(0, weight=1)

    history_lb = tk.Listbox(lb_outer, height=4, activestyle='none',
                            selectmode=tk.SINGLE, font=('TkFixedFont', 9),
                            relief='solid', borderwidth=1, highlightthickness=0,
                            selectbackground='#7ba7bc', selectforeground='white',
                            bg='white')
    history_lb.grid(row=0, column=0, sticky=tk.NSEW)

    lb_scroll = ttk.Scrollbar(lb_outer, orient=tk.VERTICAL, command=history_lb.yview)
    lb_scroll.grid(row=0, column=1, sticky=tk.NS)
    history_lb.configure(yscrollcommand=lb_scroll.set)

    def refresh_history():
        history_lb.configure(state=tk.NORMAL)
        history_lb.delete(0, tk.END)
        _hdata = _load_run_history()
        if _hdata:
            for i, entry in enumerate(_hdata):
                history_lb.insert(tk.END, f'  {_run_label(entry)}')
                status = entry.get('status', '')
                if status == 'pass':
                    history_lb.itemconfig(i, foreground='#2a7a2a')
                elif status == 'fail':
                    history_lb.itemconfig(i, foreground='#c0392b')
        else:
            history_lb.insert(tk.END, '  (no previous runs yet)')
            history_lb.configure(state=tk.DISABLED)

    refresh_history()
    history_lb.bind('<Double-Button-1>', lambda _e: load_selected_run())

    hist_btn_row = ttk.Frame(hist_lf)
    hist_btn_row.grid(row=1, column=0, sticky=tk.EW, pady=(6, 0))
    ttk.Label(hist_btn_row, text='Double-click or select then click Load →',
              foreground='grey').pack(side=tk.LEFT)
    ttk.Button(hist_btn_row, text='Load Selected Run',
               command=load_selected_run).pack(side=tk.RIGHT)

    # ── Action buttons (placed in the fixed bottom bar) ───────────────
    ttk.Button(_btn_row, text='Cancel',
               command=root.destroy).pack(side=tk.RIGHT, padx=(6, 0))
    ttk.Button(_btn_row, text='▶  Run Analysis',
               command=run, default='active').pack(side=tk.RIGHT)

    root.mainloop()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='MSD 4PL Analysis Tool')
    parser.add_argument('--msd', required=False, default=None, help='MSD .txt data file')
    parser.add_argument('--platemap', required=False, default=None, help='Plate map CSV (grid format)')
    parser.add_argument('--output', default='msd_4pl_results.xlsx', help='Output Excel file')
    parser.add_argument('--spots', type=int, choices=[1, 4, 10], default=None,
                        help='Override spots per well (auto-detected if omitted)')
    parser.add_argument('--units', default=None,
                        help='Optional units string to append to interpolated concentration headers')
    parser.add_argument('--cv-threshold', type=float, default=None,
                        help='Optional %%CV threshold for All Unknowns highlight (default 25)')
    parser.add_argument('--lloq-method', choices=['current', '3xblank'], default='current',
                        help='LLOQ calculation method: current mean+10*SD or 3x blank mean')
    parser.add_argument('--dilution-factors', default=None,
                        help='Optional per-plate dilution factors as comma-separated values (e.g. 1,2,1)')
    parser.add_argument('--total-protein', default=None,
                        help='Optional total protein CSV for normalisation (External Animal Number + Tissue Type)')
    parser.add_argument('--gui', action='store_true', help='Launch interactive file picker dialogs')
    parser.add_argument('--rerun', action='store_true', help='Rerun the last analysis with saved parameters')
    args = parser.parse_args()

    if args.rerun:
        history = _load_run_history()
        if not history:
            print("No previous run found. Use --msd and --platemap or --gui.")
            sys.exit(1)
        last_args = history[0]
        args.msd = last_args.get('msd')
        args.platemap = last_args.get('platemap')
        args.output = last_args.get('output', 'msd_4pl_results.xlsx')
        args.spots = last_args.get('spots')
        args.units = last_args.get('units')
        args.cv_threshold = last_args.get('cv_threshold', 25)
        args.lloq_method = last_args.get('lloq_method', 'current')
        args.dilution_factors = last_args.get('dilution_factors')
        args.total_protein = last_args.get('total_protein')
        args.gui = False
        print(f"Rerunning: {_run_label(last_args)}")
        print(f"  MSD: {args.msd}")
        print(f"  Plate map: {args.platemap}")
        print(f"  Output: {args.output}")

    # No args or --gui → open GUI (default when double-clicked as .app)
    if args.gui or (not args.msd and not args.platemap and not args.rerun):
        run_interactive()
    elif args.msd and args.platemap:
        run_analysis(args.msd, args.platemap, args.output, args.spots, args.units, args.cv_threshold, args.dilution_factors, args.lloq_method, args.total_protein)
    else:
        print("Error: provide both --msd and --platemap, or use --gui for interactive mode.")
        parser.print_help()

